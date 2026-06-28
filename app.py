import json
import math
import re
import calendar
import argparse
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
import tempfile
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    import streamlit as st
except ImportError:
    st = None
try:
    import plotly.express as px
except ImportError:
    px = None

# ==========================================
# CONSTANTS & CONFIGURATION
# ==========================================
APP_DIR = Path(__file__).resolve().parent
WORKBOOK = Path.home() / "Downloads" / "T&C_Resource_Planning_Report.xlsx"

LINES = ["DTL", "JRL", "CRL", "RTS"]
MONTH_MAP = {m.upper(): i for i, m in enumerate(calendar.month_abbr) if m}
MONTH_MAP.update({m.upper(): i for i, m in enumerate(calendar.month_name) if m})
LINE_COLORS = {
    "DTL": "#2563eb",
    "JRL": "#00897b",
    "CRL": "#f2c94c",
    "RTS": "#111827",
}

SUMMARY_ROLE_NAMES = {
    "TOTAL",
    "ATC/ATS",
    "ATS",
    "SIG",
    "ATC",
    "COMMS",
    "SUBCON",
    "TRAIN",
    "TRAIN KR",
    "CSF",
}

CORE_TEAM_BASIS = [
    {
        "testType": "Test Manager / Coordinator",
        "rolesRequired": "T&C Manager, T&C Coordinator",
        "siemensPax": 2,
        "subconPax": 0,
    },
    {
        "testType": "Procedure Preparation / Procedure Preparation TR",
        "rolesRequired": "ATC T&C Engineer 1, ATS T&C Engineer 2, Comms T&C Engineer 3, Sig T&C Engineer 4",
        "siemensPax": 4,
        "subconPax": 0,
    },
    {
        "testType": "Power Up and Set to Work",
        "rolesRequired": "Sig T&C Engineer 1, Comms T&C Engineer 2, ATS T&C Engineer 3, Subcon 1",
        "siemensPax": 3,
        "subconPax": 1,
    },
    {
        "testType": "Correspondence Test",
        "rolesRequired": "Sig T&C Engineer 1, Sig T&C Engineer 2, Subcon 1, Subcon 2",
        "siemensPax": 2,
        "subconPax": 2,
    },
    {
        "testType": "Through Test",
        "rolesRequired": "Sig T&C Engineer 1, Sig T&C Engineer 2, Subcon 1, Subcon 2",
        "siemensPax": 2,
        "subconPax": 2,
    },
    {
        "testType": "PAT",
        "rolesRequired": "Sig T&C Engineer 1, Sig T&C Engineer 2, Subcon 1, Subcon 2",
        "siemensPax": 2,
        "subconPax": 2,
    },
    {
        "testType": "Mainline Testing - STT / MTT / SIT / SAT",
        "rolesRequired": "ATC T&C Engineer 1, ATS T&C Engineer 2, Subcon 1, Subcon 2",
        "siemensPax": 2,
        "subconPax": 2,
    },
    {
        "testType": "Train Routine Test",
        "rolesRequired": "Train Engineer 1, Train Engineer 2, Subcon 1, Subcon 2",
        "siemensPax": 2,
        "subconPax": 2,
    },
    {
        "testType": "Train Routine Test KR",
        "rolesRequired": "Train T&C Engineer 1 KR, Train T&C Engineer 2 KR",
        "siemensPax": 2,
        "subconPax": 0,
    },
    {
        "testType": "Train Routine Testing Offsite",
        "rolesRequired": "Train T&C Engineer 1, Train T&C Engineer 2",
        "siemensPax": 2,
        "subconPax": 0,
    },
    {
        "testType": "ATC Dynamic",
        "rolesRequired": "ATC T&C Engineer 1, ATC T&C Engineer 2, ATC T&C Engineer 3, ATC T&C Engineer 4, Subcon 1 to Subcon 8",
        "siemensPax": 4,
        "subconPax": 8,
    },
    {
        "testType": "Test Running",
        "rolesRequired": "ATC/ATS T&C Engineer 1, 2, 3, 4 + same backup team, Subcon 1, 2 + same backup team",
        "siemensPax": 8,
        "subconPax": 4,
    },
    {
        "testType": "JRL CSF",
        "rolesRequired": "CSF T&C Engineer 1, CSF T&C Engineer 2",
        "siemensPax": 2,
        "subconPax": 0,
    },
    {
        "testType": "Support for Troubleshooting",
        "rolesRequired": "ATC Design Engineer, ATS Design Engineer, IXL Design Engineer, Comms Design Engineer",
        "siemensPax": 4,
        "subconPax": 0,
    },
    {
        "testType": "OSIT Meetings / Coordination",
        "rolesRequired": "OSIT Manager, T&C Coordinator, Mainline T&C Manager",
        "siemensPax": 3,
        "subconPax": 0,
    },
    {
        "testType": "Pre-OSIT",
        "rolesRequired": "T&C Engineer 1, T&C Engineer 2, RCS/Network Engineer, ATS T&C Engineer",
        "siemensPax": 4,
        "subconPax": 0,
    },
    {
        "testType": "OSIT",
        "rolesRequired": "ATS T&C Engineer 1, ATC T&C Engineer 2, ATS T&C Engineer 3, ATC T&C Engineer 4 + same backup team, Subcon 1, 2 + same backup team",
        "siemensPax": 8,
        "subconPax": 4,
    },
    {
        "testType": "OSIT Test Procedure Preparation",
        "rolesRequired": "ATC T&C Engineer 1, ATC T&C Engineer 2, ATS T&C Engineer 3, SIG T&C Engineer 4",
        "siemensPax": 4,
        "subconPax": 0,
    },
]

for core_row in CORE_TEAM_BASIS:
    core_row["totalPax"] = core_row["siemensPax"] + core_row["subconPax"]

# ==========================================
# BACKEND DATA PROCESSING LOGIC
# ==========================================
def clean(value):
    if pd.isna(value):
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)

def is_tbc_name(value):
    text = clean(value).upper()
    return text in {"", "TBC", "TBD", "NA", "N/A"} or "TBC" in text

def is_summary_role(row):
    role = clean(row.get("Role", "")).upper()
    if role == "TOTAL":
        return True
    if role in SUMMARY_ROLE_NAMES and is_tbc_name(row.get("Assigned Name", "")):
        return True
    return False

def number(value, default=0.0):
    if pd.isna(value):
        return default
    try:
        return float(value)
    except Exception:
        return default

def month_label(value):
    if pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%b %y")

def month_iso(value):
    if pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%Y-%m-%d")

def week_label(value):
    if pd.isna(value):
        return ""
    iso = pd.to_datetime(value).isocalendar()
    return f"{int(iso.year)}-W{int(iso.week):02d}"

def period_label(value, period_type="Month"):
    if clean(period_type).upper() == "WEEK":
        return week_label(value)
    return month_label(value)

def date_label(value):
    if pd.isna(value):
        return ""
    return pd.to_datetime(value).strftime("%d %b %y")

def normalise_key(value):
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())

def role_key(value):
    text = clean(value).upper()
    if not text:
        return ""
    if "SUBCON" in text or re.search(r"\bSC\b", text):
        return "SUBCON"
    if "COMMS" in text or "NETWORK" in text or "RCS" in text:
        return "COMMS T&C ENGINEER"
    if "TRAIN" in text:
        return "TRAIN T&C ENGINEER"
    if "CSF" in text:
        return "CSF T&C ENGINEER"
    if "SIG" in text or "IXL" in text:
        return "SIG T&C ENGINEER"
    if "ATS" in text:
        return "ATS T&C ENGINEER"
    if "ATC" in text:
        return "ATC T&C ENGINEER"
    if "MANAGER" in text:
        return "MANAGER"
    if "COORDINATOR" in text:
        return "COORDINATOR"
    if "DESIGN" in text:
        return "DESIGN ENGINEER"
    text = re.sub(r"\b\d+\b", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def role_keys(value):
    text = clean(value).upper()
    if not text:
        return []
    keys = set()
    if "SUBCON" in text or re.search(r"\bSC\b", text):
        keys.add("SUBCON")
    if "COMMS" in text or "NETWORK" in text or "RCS" in text:
        keys.add("COMMS T&C ENGINEER")
    if "TRAIN" in text:
        keys.add("TRAIN T&C ENGINEER")
    if "CSF" in text:
        keys.add("CSF T&C ENGINEER")
    if "SIG" in text or "IXL" in text:
        keys.add("SIG T&C ENGINEER")
    if "ATS" in text:
        keys.add("ATS T&C ENGINEER")
    if "ATC" in text:
        keys.add("ATC T&C ENGINEER")
    if "MANAGER" in text:
        keys.add("MANAGER")
    if "COORDINATOR" in text:
        keys.add("COORDINATOR")
    if "DESIGN" in text:
        keys.add("DESIGN ENGINEER")
    if not keys:
        single = role_key(value)
        if single:
            keys.add(single)
    return sorted(keys)

def preferred_company_for_role(value):
    keys = set(role_keys(value))
    if "SUBCON" in keys:
        return "SC"
    return "Siemens" if keys else ""

def normalise_company_type(value):
    text = clean(value).upper()
    if text in {"SC", "SUBCON", "SUBCONTRACTOR", "SUB-CON", "SUB CONTRACTOR"}:
        return "SC"
    if text == "SIEMENS":
        return "Siemens"
    if text in {"", "TBC"}:
        return "TBC" if text == "TBC" else "Unknown"
    return clean(value)

def classify_test_type(activity):
    text = clean(activity).lower()
    if not text or "general / unassigned" in text:
        return "Test Manager / Coordinator"
    if "correspondence" in text:
        return "Correspondence Test"
    if "power up" in text or "set to work" in text:
        return "Power Up and Set to Work"
    if "through test" in text:
        return "Through Test"
    if "atc dynamic" in text:
        return "ATC Dynamic"
    if "mainline testing" in text or "stt" in text or "mtt" in text or "sit" in text or "sat" in text:
        return "Mainline Testing - STT / MTT / SIT / SAT"
    if "test running" in text:
        return "Test Running"
    if "train routine test (kr)" in text or "train routine testing (kr)" in text:
        return "Train Routine Test KR"
    if "train routine testing (offsite)" in text or "offsite" in text:
        return "Train Routine Testing Offsite"
    if "train routine" in text:
        return "Train Routine Test"
    if "osit test procedure preparation" in text:
        return "OSIT Test Procedure Preparation"
    if "procedure preparation" in text or "procedure preparation" in text:
        return "Procedure Preparation / Procedure Preparation TR"
    if "osit meetings" in text or "coordination" in text:
        return "OSIT Meetings / Coordination"
    if "pre-osit" in text:
        return "Pre-OSIT"
    if "osit" in text:
        return "OSIT"
    if re.search(r"\bitc\b", text):
        return "ITC Test"
    if "migration" in text:
        return "Migration Test"
    if text == "pat" or " pat" in text:
        return "PAT"
    if "troubleshooting" in text:
        return "Support for Troubleshooting"
    if "manager" in text or "coordinator" in text:
        return "Test Manager / Coordinator"
    if "csf" in text:
        return "JRL CSF"
    if "pending" in text:
        return "Pending / Carryover Test"
    return "Other"

def classify_discipline(role):
    text = clean(role).lower()
    if not text or "item level demand" in text:
        return "Demand Placeholder"
    if "backup" in text:
        return "Backup"
    if "subcon" in text or re.search(r"\bsc\b", text):
        return "Subcon"
    if "design" in text:
        return "Design Support"
    if "manager" in text or "coordinator" in text:
        return "Management"
    if "train" in text:
        return "Train"
    if "comms" in text or "network" in text or "rcs" in text:
        return "Comms / Network"
    if "sig" in text or "ixl" in text:
        return "Signalling"
    if "atc" in text or "ats" in text:
        return "ATC / ATS"
    if "engineer" in text:
        return "Engineering"
    return "Other"

def parse_requirement(requirement):
    text = clean(requirement)
    lower = text.lower().replace(" ", "")
    pax = None
    pax_match = re.search(r"(\d+(?:\.\d+)?)\s*pax", text, flags=re.I)
    if pax_match:
        pax = float(pax_match.group(1))

    shifts = 1
    shift_match = re.search(r"(\d+)\s*shifts?", text, flags=re.I)
    if shift_match:
        shifts = int(shift_match.group(1))

    days = None
    day_match = re.search(r"(\d+(?:\.\d+)?)\s*days?", text, flags=re.I)
    if day_match:
        days = float(day_match.group(1))

    if "/night/slot" in lower:
        unit = "night slot"
    elif "/week" in lower:
        unit = "week"
    elif "train" in lower:
        unit = "train"
    else:
        unit = "test window"

    core_floor = pax * shifts if pax is not None else None
    return {
        "text": text,
        "pax": pax,
        "shifts": shifts,
        "days": days,
        "unit": unit,
        "coreFloor": core_floor,
    }

def to_number(value):
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    text = clean(value)
    if text in {"", "-", "–", "—"}:
        return None
    try:
        return float(text.replace(",", ""))
    except Exception:
        return None

def parse_month(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.month
    text = clean(value).upper()
    if not text:
        return None
    return MONTH_MAP.get(text[:3])

def parse_year(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, (int, float)):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    match = re.search(r"(19\d{2}|20\d{2})", clean(value))
    return int(match.group(1)) if match else None

def is_item_number(value):
    if pd.isna(value):
        return False
    if isinstance(value, (int, float)):
        return not pd.isna(value)
    return bool(re.fullmatch(r"\d+(\.\d+)?", clean(value)))

def normalise_line(value):
    text = clean(value).upper().replace(" ", "")
    return text if text in LINES else ""

def get_cell(df, row, col):
    if row >= df.shape[0] or col >= df.shape[1]:
        return None
    return df.iat[row, col]

def get_month_columns(tc_df):
    month_cols = []
    current_year = None
    for col in range(4, tc_df.shape[1]):
        year_here = parse_year(get_cell(tc_df, 0, col))
        if year_here:
            current_year = year_here
        month_num = parse_month(get_cell(tc_df, 1, col))
        if current_year and month_num:
            month_date = pd.Timestamp(current_year, month_num, 1)
            month_cols.append(
                {
                    "col_index": col,
                    "month": month_date,
                    "month_label": month_date.strftime("%b %y"),
                }
            )
    return month_cols

def parse_tc_activities(tc_df):
    month_cols = get_month_columns(tc_df)
    if not month_cols:
        raise ValueError(
            "Cannot find month headers. Expected years on row 1 and months on row 2 from column E onwards."
        )

    resource_rows = []
    monthly_records = []
    current_line = ""
    current_item_no = ""
    current_item = ""
    current_requirement = ""
    current_parent_values = {}

    for r in range(2, tc_df.shape[0]):
        a = get_cell(tc_df, r, 0)
        b = clean(get_cell(tc_df, r, 1))
        c = clean(get_cell(tc_df, r, 2))
        d = clean(get_cell(tc_df, r, 3))

        line = normalise_line(a)
        if line:
            current_line = line
            current_item_no = ""
            current_item = ""
            current_requirement = ""
            current_parent_values = {}
            continue

        if not b and not c and not d:
            continue

        numbered_item = is_item_number(a)
        is_header = numbered_item or (not numbered_item and b and d and not c)

        if is_header:
            current_item_no = clean(a) if pd.notna(a) else ""
            current_item = b
            current_requirement = d
            current_parent_values = {}
            for m in month_cols:
                v = to_number(get_cell(tc_df, r, m["col_index"]))
                if v is not None:
                    current_parent_values[m["col_index"]] = v

            has_children = False
            for curr_r in range(r + 1, tc_df.shape[0]):
                nxt_a = get_cell(tc_df, curr_r, 0)
                nxt_b = clean(get_cell(tc_df, curr_r, 1))
                if normalise_line(nxt_a) or is_item_number(nxt_a):
                    break
                if nxt_b:
                    has_children = True
                    break

            if has_children:
                continue

            role = "Item level demand"
            assigned_name = "TBC"
            row_type = "Item level demand"
            row_requirement = d
            values_by_month = current_parent_values
        else:
            if not current_item:
                current_item = "General / Unassigned item"
                current_item_no = ""
                current_requirement = d

            role = b
            assigned_name = c if c else "TBC"
            row_type = "Role row"
            row_requirement = d if d else current_requirement
            values_by_month = {}
            for m in month_cols:
                v = to_number(get_cell(tc_df, r, m["col_index"]))
                if v is not None:
                    values_by_month[m["col_index"]] = v

            if not values_by_month:
                values_by_month = current_parent_values.copy()

        if not values_by_month:
            continue

        source_row = r + 1
        row_id = f"R{source_row}_{current_line}_{current_item_no}_{role}_{assigned_name}"
        load_list = [(m, values_by_month[m["col_index"]]) for m in month_cols if m["col_index"] in values_by_month]
        start_month = min([x[0]["month"] for x in load_list], default=pd.NaT)
        finish_month = max([x[0]["month"] for x in load_list], default=pd.NaT)

        resource_rows.append(
            {
                "Row ID": row_id,
                "Source Row": source_row,
                "Line": current_line,
                "Item No": current_item_no,
                "Activity / Item": current_item,
                "Item Requirement": row_requirement,
                "Role": role,
                "Current Name": assigned_name,
                "Assigned Name": assigned_name,
                "Row Type": row_type,
                "Start Month": start_month,
                "Finish Month": finish_month,
            }
        )

        for m, load in load_list:
            monthly_records.append(
                {
                    "Row ID": row_id,
                    "Source Row": source_row,
                    "Line": current_line,
                    "Item No": current_item_no,
                    "Activity / Item": current_item,
                    "Item Requirement": row_requirement,
                    "Role": role,
                    "Current Name": assigned_name,
                    "Month": m["month"],
                    "Month Label": m["month_label"],
                    "Load": load,
                }
            )

    resources = pd.DataFrame(resource_rows)
    monthly = pd.DataFrame(monthly_records)
    if resources.empty:
        raise ValueError("No resource rows found from T&C Activities.")
    return resources, monthly

def parse_manpower(mp_df):
    records = []
    for start_col in [0, 3, 6, 9]:
        line = clean(get_cell(mp_df, 1, start_col)).upper()
        if not line:
            continue
        for r in range(2, mp_df.shape[0]):
            name = clean(get_cell(mp_df, r, start_col))
            company = clean(get_cell(mp_df, r, start_col + 1))
            if not name:
                continue
            company_upper = company.upper()
            if company_upper in {"SC", "SUBCON", "SUBCONTRACTOR"}:
                company = "SC"
            elif company_upper == "SIEMENS":
                company = "Siemens"
            elif not company:
                company = "Unknown"
            records.append({"Line": line, "Name": name, "Company": company, "Role": "", "Max Weekly Load": 1.0})

    manpower = pd.DataFrame(records)
    if manpower.empty:
        manpower = pd.DataFrame(columns=["Line", "Name", "Company", "Role", "Max Weekly Load"])
    manpower["Name Key"] = manpower["Name"].astype(str).str.strip().str.upper()
    return manpower

TABLE_ALIASES = {
    "test_type": ["Test Type", "TestType", "Activity Type", "Testing Type"],
    "role_required": ["Role Required", "Roles Required", "Role", "Required Role"],
    "company_type": ["Company Type", "Company", "Resource Type"],
    "line": ["Line", "Project", "Project Line"],
    "start_date": ["Start Date", "Start", "StartDate", "From"],
    "end_date": ["End Date", "Finish Date", "Finish", "EndDate", "To"],
    "hours_day": ["Hours/Day", "Hours Per Day", "Hours Day", "Hrs/Day", "Hours"],
    "person": ["Person", "Name", "Resource", "Resource Name"],
    "role_discipline": ["Role / Discipline", "Role Discipline", "Discipline", "Role"],
    "max_weekly_load": ["Max Weekly Load", "Max Load", "Weekly Capacity", "Capacity"],
    "assigned_person": ["Assigned Person", "Assigned Name", "Person", "Name"],
}

def find_column(columns, aliases):
    lookup = {normalise_key(col): col for col in columns}
    for alias in aliases:
        key = normalise_key(alias)
        if key in lookup:
            return lookup[key]
    for key, col in lookup.items():
        if any(normalise_key(alias) in key for alias in aliases):
            return col
    return None

def read_table_sheet(workbook_path, sheet_name, required_fields):
    raw = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    if raw.empty:
        return pd.DataFrame()
    max_scan = min(12, len(raw))
    for header_idx in range(max_scan):
        header_values = [clean(value) for value in raw.iloc[header_idx].tolist()]
        columns = [value if value else f"Column {idx + 1}" for idx, value in enumerate(header_values)]
        mapping = {}
        for field in required_fields:
            col = find_column(columns, TABLE_ALIASES[field])
            if col is None:
                break
            mapping[field] = col
        if len(mapping) == len(required_fields):
            table = raw.iloc[header_idx + 1 :].copy()
            table.columns = columns
            table = table.dropna(how="all")
            return table.reset_index(drop=True)
    return pd.DataFrame()

def find_table_sheet(sheet_names, candidates):
    for sheet_name in sheet_names:
        key = normalise_key(sheet_name)
        if any(candidate in key for candidate in candidates):
            return sheet_name
    return None

def table_value(row, columns, field, default=""):
    col = find_column(columns, TABLE_ALIASES[field])
    if col is None:
        return default
    value = row.get(col, default)
    if pd.isna(value):
        return default
    return value

def coerce_date(value):
    if pd.isna(value):
        return pd.NaT
    try:
        parsed = pd.to_datetime(value, dayfirst=True)
    except Exception:
        parsed = pd.NaT
    return parsed

def split_role_list(value):
    roles = []
    for part in re.split(r",", clean(value)):
        part = clean(part)
        if not part:
            continue
        span = re.match(r"(.+?)\s+(\d+)\s+to\s+(?:.+?\s+)?(\d+)$", part, flags=re.I)
        if span:
            prefix = clean(span.group(1))
            start = int(span.group(2))
            end = int(span.group(3))
            for idx in range(start, end + 1):
                roles.append(f"{prefix} {idx}")
        else:
            roles.append(part)
    return roles

def core_requirement_master():
    records = []
    for item in CORE_TEAM_BASIS:
        for role in split_role_list(item["rolesRequired"]):
            records.append(
                {
                    "Test Type": item["testType"],
                    "Role Required": role,
                    "Company Type": "SC" if "SUBCON" in role.upper() else "Siemens",
                }
            )
    return records

def iter_week_loads(start_date, end_date, hours_per_day):
    start = coerce_date(start_date)
    end = coerce_date(end_date)
    if pd.isna(start) or pd.isna(end):
        return []
    if end < start:
        start, end = end, start
    hours = number(hours_per_day, 8.0) or 8.0
    weekly_hours = defaultdict(float)
    current = start.normalize()
    end = end.normalize()
    while current <= end:
        if current.weekday() < 5:
            week_start = current - timedelta(days=current.weekday())
            weekly_hours[week_start] += hours
        current += timedelta(days=1)
    return [
        {
            "week_start": week_start,
            "week_label": week_label(week_start),
            "load": worked_hours / 40.0,
        }
        for week_start, worked_hours in sorted(weekly_hours.items())
        if worked_hours > 0
    ]

def parse_weekly_planning_workbook(workbook_path, xl):
    sheet_names = xl.sheet_names
    schedule_sheet = find_table_sheet(sheet_names, ["activityschedule", "schedule", "activities"])
    people_sheet = find_table_sheet(sheet_names, ["resourcelist", "peoplelist", "resources", "people", "manpower"])
    assignment_sheet = find_table_sheet(sheet_names, ["assignmenttable", "assignments", "allocation"])
    requirement_sheet = find_table_sheet(sheet_names, ["testtyperequirementmaster", "requirementmaster", "requirements"])

    if not (schedule_sheet and people_sheet and assignment_sheet):
        return None

    schedule = read_table_sheet(workbook_path, schedule_sheet, ["line", "test_type", "start_date", "end_date"])
    people = read_table_sheet(workbook_path, people_sheet, ["person", "role_discipline", "company_type"])
    assignments = read_table_sheet(workbook_path, assignment_sheet, ["line", "test_type", "role_required"])
    if schedule.empty or people.empty or assignments.empty:
        return None

    if requirement_sheet:
        requirements = read_table_sheet(workbook_path, requirement_sheet, ["test_type", "role_required", "company_type"])
    else:
        requirements = pd.DataFrame(core_requirement_master())

    requirement_records = []
    for _, row in requirements.iterrows():
        columns = requirements.columns
        test_type = clean(table_value(row, columns, "test_type"))
        role = clean(table_value(row, columns, "role_required"))
        if not test_type or not role:
            continue
        requirement_records.append(
            {
                "test_type": test_type,
                "role": role,
                "company": normalise_company_type(table_value(row, columns, "company_type", "Siemens")),
            }
        )
    requirements_by_test = defaultdict(list)
    for item in requirement_records:
        requirements_by_test[normalise_key(item["test_type"])].append(item)

    people_records = []
    for _, row in people.iterrows():
        columns = people.columns
        name = clean(table_value(row, columns, "person"))
        if not name:
            continue
        max_load = to_number(table_value(row, columns, "max_weekly_load", 1.0))
        people_records.append(
            {
                "Line": clean(table_value(row, columns, "line", "ALL")).upper() or "ALL",
                "Name": name,
                "Company": normalise_company_type(table_value(row, columns, "company_type", "Siemens")),
                "Role": clean(table_value(row, columns, "role_discipline")),
                "Max Weekly Load": max_load if max_load is not None else 1.0,
            }
        )
    manpower = pd.DataFrame(people_records)
    if manpower.empty:
        manpower = pd.DataFrame(columns=["Line", "Name", "Company", "Role", "Max Weekly Load"])
    manpower["Name Key"] = manpower["Name"].astype(str).str.strip().str.upper()
    name_to_company = dict(zip(manpower["Name Key"], manpower["Company"]))

    assignment_lookup = {}
    for _, row in assignments.iterrows():
        columns = assignments.columns
        line = normalise_line(table_value(row, columns, "line"))
        test_type = clean(table_value(row, columns, "test_type"))
        role = clean(table_value(row, columns, "role_required"))
        assigned = clean(table_value(row, columns, "assigned_person", "TBC")) or "TBC"
        if not test_type or not role:
            continue
        assignment_lookup[(line, normalise_key(test_type), normalise_key(role))] = assigned
        assignment_lookup[("", normalise_key(test_type), normalise_key(role))] = assigned

    resources = []
    monthly = []
    source_row = 1
    for idx, row in schedule.iterrows():
        columns = schedule.columns
        line = normalise_line(table_value(row, columns, "line"))
        test_type = clean(table_value(row, columns, "test_type"))
        if not line or not test_type:
            continue
        start_date = coerce_date(table_value(row, columns, "start_date"))
        end_date = coerce_date(table_value(row, columns, "end_date"))
        hours_day = table_value(row, columns, "hours_day", 8)
        week_loads = iter_week_loads(start_date, end_date, hours_day)
        if not week_loads:
            continue
        roles = requirements_by_test.get(normalise_key(test_type), [])
        if not roles:
            roles = [{"test_type": test_type, "role": "Item level demand", "company": "TBC"}]

        for role_idx, req in enumerate(roles, start=1):
            role = req["role"]
            assigned = assignment_lookup.get((line, normalise_key(test_type), normalise_key(role)))
            if assigned is None:
                assigned = assignment_lookup.get(("", normalise_key(test_type), normalise_key(role)), "TBC")
            assigned = clean(assigned) or "TBC"
            row_id = f"W{idx + 1}_{line}_{normalise_key(test_type)}_{role_idx}_{normalise_key(role)}"
            company = "TBC" if is_tbc_name(assigned) else name_to_company.get(assigned.upper(), req["company"])
            resources.append(
                {
                    "Row ID": row_id,
                    "Source Row": source_row,
                    "Line": line,
                    "Item No": idx + 1,
                    "Activity / Item": test_type,
                    "Item Requirement": f"{number(hours_day, 8):g}h/day",
                    "Role": role,
                    "Current Name": assigned,
                    "Assigned Name": assigned,
                    "Row Type": "Role row",
                    "Start Month": start_date,
                    "Finish Month": end_date,
                    "TBC?": "YES" if is_tbc_name(assigned) else "NO",
                    "Role + Name": f"{role} / {assigned}",
                    "Company": company,
                }
            )
            for week in week_loads:
                monthly.append(
                    {
                        "Row ID": row_id,
                        "Source Row": source_row,
                        "Line": line,
                        "Item No": idx + 1,
                        "Activity / Item": test_type,
                        "Item Requirement": f"{number(hours_day, 8):g}h/day",
                        "Role": role,
                        "Current Name": assigned,
                        "Month": week["week_start"],
                        "Month Label": week["week_label"],
                        "Period Type": "Week",
                        "Load": week["load"],
                        "Assigned Name": assigned,
                        "TBC?": "YES" if is_tbc_name(assigned) else "NO",
                        "Role + Name": f"{role} / {assigned}",
                        "Company": company,
                    }
                )
            source_row += 1

    if not resources or not monthly:
        return None
    return (
        pd.DataFrame(resources),
        pd.DataFrame(monthly),
        manpower.drop(columns=["Name Key"], errors="ignore"),
        pd.DataFrame(),
        "Weekly planning workbook",
    )

def classify_company(row, name_to_company):
    assigned = clean(row.get("Assigned Name", ""))
    role = clean(row.get("Role", ""))
    if is_tbc_name(assigned):
        return "TBC"
    key = assigned.upper()
    if key in name_to_company:
        return name_to_company[key]
    if "SUBCON" in role.upper() or assigned.upper().startswith("SUBCON") or assigned.upper() == "SC":
        return "SC"
    return "Siemens"

def apply_company_fields(resources, monthly, manpower):
    resources = resources.copy()
    monthly = monthly.copy()
    resources["Assigned Name"] = resources["Assigned Name"].apply(lambda x: clean(x) if clean(x) else "TBC")
    resources["TBC?"] = resources["Assigned Name"].apply(lambda x: "YES" if is_tbc_name(x) else "NO")
    resources["Role + Name"] = resources["Role"].astype(str) + " / " + resources["Assigned Name"].astype(str)
    name_to_company = dict(zip(manpower.get("Name Key", pd.Series(dtype=str)), manpower.get("Company", pd.Series(dtype=str))))
    resources["Company"] = resources.apply(lambda row: classify_company(row, name_to_company), axis=1)
    join_cols = resources[["Row ID", "Assigned Name", "TBC?", "Role + Name", "Company"]]
    monthly = monthly.drop(columns=[c for c in ["Assigned Name", "TBC?", "Role + Name", "Company"] if c in monthly.columns])
    monthly = monthly.merge(join_cols, on="Row ID", how="left")
    return resources, monthly

def make_person_overload(monthly, person_threshold=1.0):
    if monthly.empty or "Assigned Name" not in monthly.columns:
        return pd.DataFrame()
    valid = monthly[~monthly["Assigned Name"].apply(is_tbc_name)].copy()
    if valid.empty:
        return pd.DataFrame()
    person_load = (
        valid.groupby(["Month", "Month Label", "Assigned Name", "Company"], as_index=False)["Load"]
        .sum()
        .sort_values(["Month", "Assigned Name"])
    )
    person_load["Person Threshold"] = float(person_threshold)
    person_load["Exceeded?"] = person_load["Load"].apply(lambda x: "YES" if x > person_threshold else "NO")
    return person_load[person_load["Load"] > person_threshold].copy()

def make_issue_summary(person_load_records, issue_records):
    return {
        "personClashes": int(sum(1 for row in issue_records if row.get("status") == "Clash")),
        "tbcGaps": int(sum(1 for row in issue_records if row.get("status") == "TBC Gap")),
        "fullyLoaded": int(sum(1 for row in person_load_records if row.get("status") == "Fully Loaded")),
        "available": int(sum(1 for row in person_load_records if row.get("status") == "Available")),
        "issues": int(len(issue_records)),
    }

def load_status(load, threshold):
    if load > threshold + 1e-9:
        return "Clash"
    if abs(load - threshold) <= 1e-9:
        return "Fully Loaded"
    return "Available"

def make_people_roster(manpower, resources=None):
    people = {}

    def ensure_person(name):
        key = clean(name).upper()
        if key not in people:
            people[key] = {
                "name": clean(name),
                "role": "",
                "roleKey": "",
                "roleKeys": [],
                "company": "",
                "maxWeeklyLoad": 1.0,
                "lines": set(),
            }
        return people[key]

    if not manpower.empty:
        for _, row in manpower.iterrows():
            name = clean(row.get("Name", ""))
            if not name or is_tbc_name(name):
                continue
            max_load = to_number(row.get("Max Weekly Load", 1.0))
            person = ensure_person(name)
            role = clean(row.get("Role", ""))
            if role:
                person["role"] = role
                person["roleKey"] = role_key(role)
                person["roleKeys"] = role_keys(role)
            person["company"] = normalise_company_type(row.get("Company", ""))
            person["maxWeeklyLoad"] = max_load if max_load is not None else 1.0
            line = normalise_line(row.get("Line", "")) or clean(row.get("Line", "")).upper()
            if line and line != "ALL":
                person["lines"].add(line)

    if resources is not None and not resources.empty:
        for _, row in resources.iterrows():
            name = clean(row.get("Assigned Name", ""))
            if not name or is_tbc_name(name):
                continue
            current = ensure_person(name)
            inferred_role = clean(row.get("Role", ""))
            if inferred_role and not current.get("role"):
                current["role"] = inferred_role
                current["roleKey"] = role_key(inferred_role)
                current["roleKeys"] = role_keys(inferred_role)
            if not current.get("company"):
                current["company"] = normalise_company_type(row.get("Company", ""))
            current["maxWeeklyLoad"] = current.get("maxWeeklyLoad") or 1.0
            line = normalise_line(row.get("Line", "")) or clean(row.get("Line", "")).upper()
            if line:
                current["lines"].add(line)

    roster = []
    for person in people.values():
        lines = sorted(person.pop("lines", set()))
        person["lines"] = lines
        person["line"] = ", ".join(lines)
        person["roleKeys"] = person.get("roleKeys") or role_keys(person.get("role", ""))
        if not person.get("roleKey") and person["roleKeys"]:
            person["roleKey"] = person["roleKeys"][0]
        if not person.get("company"):
            person["company"] = "Unknown"
        roster.append(person)
    return sorted(roster, key=lambda row: row["name"].upper())

def max_load_map_from_roster(people_roster):
    return {row["name"].upper(): number(row.get("maxWeeklyLoad", 1.0), 1.0) for row in people_roster}

def suggest_issue_fix(issue, loads_by_person_period, people_roster):
    status = issue.get("status")
    role = clean(issue.get("role", ""))
    target_role_keys = set(role_keys(role))
    target_role_label = "/".join(sorted(target_role_keys)) if target_role_keys else role
    period = issue.get("month", "")
    period_label_text = issue.get("monthLabel") or issue.get("weekLabel") or period
    assigned = clean(issue.get("assigned", ""))
    load = number(issue.get("load", 0))
    threshold = number(issue.get("threshold", 1), 1)
    excess = number(issue.get("excessLoad", 0))
    issue_line = clean(issue.get("line", ""))
    issue_company = normalise_company_type(issue.get("company", ""))
    preferred_company = preferred_company_for_role(role) or ("" if issue_company in {"TBC", "Unknown"} else issue_company)

    candidates = []
    for person in people_roster:
        person_role_keys = set(person.get("roleKeys") or role_keys(person.get("role", "")))
        if target_role_keys and not (person_role_keys & target_role_keys):
            continue
        if assigned and person["name"].upper() == assigned.upper():
            continue
        current_load = loads_by_person_period.get((person["name"].upper(), period), 0.0)
        capacity = number(person.get("maxWeeklyLoad", 1.0), 1.0) - current_load
        if capacity > 1e-9:
            line_match = 1 if issue_line and issue_line in (person.get("lines") or []) else 0
            company_match = 1 if preferred_company and normalise_company_type(person.get("company", "")) == preferred_company else 0
            matched_roles = sorted(person_role_keys & target_role_keys) if target_role_keys else sorted(person_role_keys)
            candidates.append((company_match, line_match, capacity, person, current_load, matched_roles))
    candidates.sort(key=lambda item: (-item[0], -item[1], -item[2], item[3]["name"]))

    needed = excess if status == "Clash" else load
    if candidates:
        _, _, capacity, person, _, matched_roles = candidates[0]
        qualifier = []
        if matched_roles:
            qualifier.append("/".join(matched_roles))
        if person.get("line"):
            qualifier.append(person["line"])
        qualifier_text = f" ({', '.join(qualifier)})" if qualifier else ""
        if capacity >= needed - 1e-9:
            action = "Assign" if status == "TBC Gap" else "Move"
            return f"{action} {rounded(needed, 2)} load to {person['name']}{qualifier_text}; {rounded(capacity, 2)} capacity available in {period_label_text}."
        return f"Split: {person['name']}{qualifier_text} can take {rounded(capacity, 2)} in {period_label_text}; add another {target_role_label or 'matching'} resource or shift {rounded(max(needed - capacity, 0), 2)} load."

    if status == "Clash":
        return f"No spare {target_role_label or 'matching role'} capacity found in {period_label_text}; add {rounded(excess, 2)} load or move one activity to another period."
    if status == "TBC Gap":
        return f"No spare {target_role_label or 'matching role'} capacity found in {period_label_text}; assign a new person or shift {rounded(load, 2)} load."
    if status == "Fully Loaded":
        return "OK - full weekly load."
    return f"OK - {rounded(max(threshold - load, 0), 2)} capacity remaining."

def make_person_load_tables(monthly, people_roster=None, person_threshold=1.0):
    if monthly.empty or "Assigned Name" not in monthly.columns:
        return [], [], [], make_issue_summary([], [])
    people_roster = people_roster or []
    max_load_by_name = max_load_map_from_roster(people_roster)

    valid = monthly[~monthly["Assigned Name"].apply(is_tbc_name)].copy()
    load_records = []
    clash_records = []
    issue_records = []
    loads_by_person_period = {}

    if not valid.empty:
        agg = (
            valid.groupby(["Month", "Month Label", "Assigned Name", "Company"], as_index=False)
            .agg(
                load=("Load", "sum"),
                row_ids=("Row ID", lambda s: sorted(set(clean(v) for v in s if clean(v)))),
                roles=("Role", lambda s: sorted(set(clean(v) for v in s if clean(v)))),
            )
            .sort_values(["Month", "Assigned Name"])
        )
        for _, row in agg.iterrows():
            assigned = clean(row["Assigned Name"])
            threshold = max_load_by_name.get(assigned.upper(), person_threshold)
            status = load_status(number(row["load"]), threshold)
            period_iso = month_iso(row["Month"])
            item = {
                "month": period_iso,
                "monthLabel": clean(row["Month Label"]),
                "week": period_iso,
                "weekLabel": clean(row["Month Label"]),
                "assigned": assigned,
                "company": clean(row["Company"]),
                "role": clean(", ".join(row["roles"][:3])) if isinstance(row["roles"], list) else "",
                "load": rounded(row["load"], 2),
                "threshold": rounded(threshold, 2),
                "status": status,
                "excessLoad": rounded(max(number(row["load"]) - threshold, 0), 2),
                "excessDays": rounded(max(number(row["load"]) - threshold, 0) * 5, 2),
                "rowIds": row["row_ids"] if isinstance(row["row_ids"], list) else [],
                "clash": status == "Clash",
            }
            loads_by_person_period[(assigned.upper(), period_iso)] = number(row["load"])
            item["suggestedFix"] = ""
            load_records.append(item)
            if status == "Clash":
                clash_records.append(item.copy())
                issue_records.append(item.copy())

    tbc = monthly[monthly["Assigned Name"].apply(is_tbc_name)].copy()
    if not tbc.empty:
        for _, row in tbc.iterrows():
            item = {
                "month": month_iso(row["Month"]),
                "monthLabel": clean(row["Month Label"]),
                "week": month_iso(row["Month"]),
                "weekLabel": clean(row["Month Label"]),
                "assigned": "TBC",
                "company": "TBC",
                "role": clean(row.get("Role", "")),
                "line": clean(row.get("Line", "")),
                "testType": clean(row.get("Test Type", classify_test_type(row.get("Activity / Item", "")))),
                "activity": clean(row.get("Activity / Item", "")),
                "load": rounded(row.get("Load", 0), 2),
                "threshold": person_threshold,
                "status": "TBC Gap",
                "excessLoad": rounded(row.get("Load", 0), 2),
                "excessDays": rounded(number(row.get("Load", 0)) * 5, 2),
                "rowIds": [clean(row.get("Row ID", ""))],
                "clash": False,
                "suggestedFix": "",
            }
            load_records.append(item.copy())
            issue_records.append(item)

    for item in issue_records:
        item["suggestedFix"] = suggest_issue_fix(item, loads_by_person_period, people_roster)
    for item in load_records:
        if item.get("status") in {"Available", "Fully Loaded"}:
            item["suggestedFix"] = suggest_issue_fix(item, loads_by_person_period, people_roster)

    load_records.sort(key=lambda row: (row.get("month", ""), row.get("assigned", ""), row.get("status", "")))
    clash_records.sort(key=lambda row: (row.get("month", ""), -number(row.get("load", 0)), row.get("assigned", "")))
    issue_records.sort(key=lambda row: (row.get("month", ""), row.get("status", ""), -number(row.get("excessLoad", 0))))
    return load_records, clash_records, issue_records, make_issue_summary(load_records, issue_records)

def load_source_tables(workbook_path):
    workbook_path = Path(workbook_path)
    xl = pd.ExcelFile(workbook_path, engine="openpyxl")
    sheets = set(xl.sheet_names)

    weekly_tables = parse_weekly_planning_workbook(workbook_path, xl)
    if weekly_tables is not None:
        return weekly_tables

    if {"Assignment_Overview", "Monthly_Detail"}.issubset(sheets):
        resources = pd.read_excel(workbook_path, sheet_name="Assignment_Overview", engine="openpyxl")
        monthly = pd.read_excel(workbook_path, sheet_name="Monthly_Detail", engine="openpyxl")
        if "Manpower_Master" in sheets:
            manpower = pd.read_excel(workbook_path, sheet_name="Manpower_Master", engine="openpyxl")
        else:
            manpower = pd.DataFrame(columns=["Line", "Name", "Company"])
        if "Overload_List" in sheets:
            overload = pd.read_excel(workbook_path, sheet_name="Overload_List", engine="openpyxl")
        else:
            overload = make_person_overload(monthly)
        return resources, monthly, manpower, overload, "Generated report"

    if {"T&C Activities", "Manpower"}.issubset(sheets):
        tc_df = pd.read_excel(workbook_path, sheet_name="T&C Activities", header=None, engine="openpyxl")
        mp_df = pd.read_excel(workbook_path, sheet_name="Manpower", header=None, engine="openpyxl")
        resources, monthly = parse_tc_activities(tc_df)
        manpower = parse_manpower(mp_df)
        resources, monthly = apply_company_fields(resources, monthly, manpower)
        overload = make_person_overload(monthly)
        return resources, monthly, manpower.drop(columns=["Name Key"], errors="ignore"), overload, "Raw T&C workbook"

    raise ValueError(
        "Workbook must contain either the four planning tables (Requirement Master, Activity Schedule, People List, Assignment Table), Assignment_Overview + Monthly_Detail, or T&C Activities + Manpower."
    )

def median(values):
    values = sorted([v for v in values if v is not None])
    if not values:
        return None
    n = len(values)
    mid = n // 2
    if n % 2:
        return float(values[mid])
    return float((values[mid - 1] + values[mid]) / 2)

def percentile(values, pct):
    values = sorted([v for v in values if v is not None])
    if not values:
        return None
    if len(values) == 1:
        return float(values[0])
    pos = (len(values) - 1) * pct
    low = math.floor(pos)
    high = math.ceil(pos)
    if low == high:
        return float(values[int(pos)])
    return float(values[low] * (high - pos) + values[high] * (pos - low))

def mode_number(values):
    rounded_vals = [round(v, 2) for v in values if v is not None]
    if not rounded_vals:
        return None
    return float(Counter(rounded_vals).most_common(1)[0][0])

def as_records(df):
    records = []
    for row in df.to_dict(orient="records"):
        item = {}
        for key, value in row.items():
            if pd.isna(value):
                item[key] = None
            elif isinstance(value, pd.Timestamp):
                item[key] = value.strftime("%Y-%m-%d")
            elif isinstance(value, (int, float)):
                item[key] = float(value)
            else:
                item[key] = str(value)
        records.append(item)
    return records

def rounded(value, digits=1):
    if value is None:
        return None
    return round(float(value), digits)

def empty_data(source_label="No workbook loaded"):
    line_metrics = []
    for line in LINES:
        line_metrics.append(
            {
                "line": line,
                "peakDemand": 0,
                "peakMonth": "",
                "peakMonthISO": "",
                "available": 0,
                "gap": 0,
                "tbcRows": 0,
                "activityRows": 0,
                "drivers": [],
                "why": f"Upload an Excel workbook to calculate {line} demand.",
            }
        )

    return {
        "summary": {
            "source": source_label,
            "sourceMode": "Waiting for upload",
            "resourceRows": 0,
            "monthlyRows": 0,
            "excludedResourceRows": 0,
            "excludedMonthlyRows": 0,
            "totalManpower": 0,
            "tbcRows": 0,
            "overloadPersonMonths": 0,
            "peakDemand": 0,
            "peakMonth": "",
            "testTypes": 0,
            "lines": LINES,
        },
        "lineColors": LINE_COLORS,
        "months": [],
        "lines": LINES,
        "manpower": {line: {"total": 0, "Siemens": 0, "SC": 0} for line in LINES},
        "lineMetrics": line_metrics,
        "demandRows": [],
        "companyRows": [],
        "testBasis": CORE_TEAM_BASIS,
        "testLineMatrix": [],
        "testTypes": [row["testType"] for row in CORE_TEAM_BASIS],
        "instances": [],
        "detailRows": [],
        "personLoads": [],
        "clashRows": [],
        "issueRows": [],
        "issueSummary": make_issue_summary([], []),
        "peopleRoster": [],
        "peopleOptions": [],
        "periodType": "Month",
        "viewMode": "Month",
        "viewModes": ["Month", "Week"],
        "weeklyExact": False,
        "demandRowsByView": {"Month": [], "Week": []},
        "periodsByView": {"Month": [], "Week": []},
    }

def build_data(workbook_path=WORKBOOK, source_label=None):
    resources, monthly, manpower, overload, source_mode = load_source_tables(workbook_path)
    source_label = source_label or Path(workbook_path).name

    resources["Excluded Summary Row"] = resources.apply(is_summary_role, axis=1)
    monthly["Excluded Summary Row"] = monthly.apply(is_summary_role, axis=1)
    excluded_resource_rows = int(resources["Excluded Summary Row"].sum())
    excluded_monthly_rows = int(monthly["Excluded Summary Row"].sum())
    resources = resources[~resources["Excluded Summary Row"]].copy()
    monthly = monthly[~monthly["Excluded Summary Row"]].copy()
    if "Period Type" not in monthly.columns:
        monthly["Period Type"] = "Month"
    period_type = "Week" if (monthly["Period Type"].astype(str).str.upper() == "WEEK").any() else "Month"

    for df in [resources, monthly]:
        df["Test Type"] = df["Activity / Item"].apply(classify_test_type)
        df["Discipline"] = df["Role"].apply(classify_discipline)
        parsed = df["Item Requirement"].apply(parse_requirement)
        df["Requirement Pax"] = parsed.apply(lambda x: x["pax"])
        df["Requirement Shifts"] = parsed.apply(lambda x: x["shifts"])
        df["Core Floor"] = parsed.apply(lambda x: x["coreFloor"])
        df["Requirement Unit"] = parsed.apply(lambda x: x["unit"])

    monthly["Month ISO"] = monthly["Month"].apply(month_iso)
    monthly["Month Label"] = monthly["Month"].apply(lambda value: period_label(value, period_type))
    monthly["View Week"] = monthly["Month"]
    monthly["View Week Label"] = monthly["Month Label"] if period_type == "Week" else monthly["Month"].apply(month_label)
    monthly["View Month"] = pd.to_datetime(monthly["Month"]).dt.to_period("M").dt.to_timestamp()
    monthly["View Month Label"] = monthly["View Month"].apply(month_label)
    if period_type == "Week":
        resources["Start Label"] = resources["Start Month"].apply(date_label)
        resources["Finish Label"] = resources["Finish Month"].apply(date_label)
    else:
        resources["Start Label"] = resources["Start Month"].apply(month_label)
        resources["Finish Label"] = resources["Finish Month"].apply(month_label)
    people_roster = make_people_roster(manpower, resources)
    person_load_records, clash_records, issue_records, issue_summary = make_person_load_tables(monthly, people_roster)
    all_test_types = sorted(set(monthly["Test Type"].dropna().astype(str)) | {row["testType"] for row in CORE_TEAM_BASIS})

    demand_line = (
        monthly.pivot_table(
            index="Month",
            columns="Line",
            values="Load",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values("Month")
    )
    for line in LINES:
        if line not in demand_line.columns:
            demand_line[line] = 0.0
    demand_line = demand_line[["Month"] + LINES]
    demand_line["Total"] = demand_line[LINES].sum(axis=1)
    demand_line["Threshold"] = float(len(manpower))

    companies = ["Siemens", "SC", "TBC", "Unknown"]
    demand_company = (
        monthly.pivot_table(
            index="Month",
            columns="Company",
            values="Load",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values("Month")
    )
    for company in companies:
        if company not in demand_company.columns:
            demand_company[company] = 0.0
    demand_company = demand_company[["Month"] + companies]
    demand_company["Total"] = demand_company[companies].sum(axis=1)

    demand_line["Month ISO"] = demand_line["Month"].apply(month_iso)
    demand_line["Month Label"] = demand_line["Month"].apply(month_label)
    demand_company["Month ISO"] = demand_company["Month"].apply(month_iso)
    demand_company["Month Label"] = demand_company["Month"].apply(month_label)

    months = (
        demand_line[["Month ISO", "Month Label"]]
        .drop_duplicates()
        .sort_values("Month ISO")
        .to_dict(orient="records")
    )

    manpower_summary = {}
    for line in LINES:
        line_series = manpower["Line"].astype(str).str.upper() if "Line" in manpower.columns else pd.Series(dtype=str)
        line_mp = manpower[line_series.isin([line, "ALL", ""])]
        manpower_summary[line] = {
            "total": int(len(line_mp)),
            "Siemens": int((line_mp["Company"] == "Siemens").sum()),
            "SC": int((line_mp["Company"] == "SC").sum()),
        }

    line_metrics = []
    for line in LINES:
        series = demand_line[["Month ISO", "Month Label", line]].rename(columns={line: "Demand"})
        peak_row = series.loc[series["Demand"].idxmax()] if len(series) else None
        peak_month = peak_row["Month ISO"] if peak_row is not None else ""
        peak_label = peak_row["Month Label"] if peak_row is not None else ""
        peak_demand = number(peak_row["Demand"]) if peak_row is not None else 0
        available = manpower_summary[line]["total"]
        line_resources = resources[resources["Line"] == line]
        line_monthly = monthly[monthly["Line"] == line]
        tbc_rows = int((line_resources["TBC?"] == "YES").sum())

        peak_monthly = line_monthly[line_monthly["Month ISO"] == peak_month]
        drivers = (
            peak_monthly.groupby(["Test Type", "Activity / Item", "Item Requirement"], dropna=False)
            .agg(
                load=("Load", "sum"),
                rows=("Row ID", "nunique"),
                roles=("Role", lambda s: ", ".join(sorted(set(clean(v) for v in s if clean(v)))[:6])),
            )
            .reset_index()
            .sort_values("load", ascending=False)
            .head(5)
        )
        driver_records = [
            {
                "testType": row["Test Type"],
                "activity": row["Activity / Item"],
                "requirement": clean(row["Item Requirement"]),
                "load": rounded(row["load"]),
                "rows": int(row["rows"]),
                "roles": row["roles"],
            }
            for _, row in drivers.iterrows()
        ]

        top_driver = driver_records[0] if driver_records else None
        why = (
            f"{line} peaks at {rounded(peak_demand)} in {peak_label}, driven mainly by "
            f"{top_driver['activity']} ({top_driver['load']} load; {top_driver['requirement'] or 'no stated requirement'})."
            if top_driver
            else f"{line} has no monthly demand in the current dataset."
        )

        line_metrics.append(
            {
                "line": line,
                "peakDemand": rounded(peak_demand),
                "peakMonth": peak_label,
                "peakMonthISO": peak_month,
                "available": available,
                "gap": rounded(peak_demand - available),
                "tbcRows": tbc_rows,
                "activityRows": int(len(line_resources)),
                "drivers": driver_records,
                "why": why,
            }
        )

    instances = []
    for (line, activity, test_type), group in resources.groupby(["Line", "Activity / Item", "Test Type"], dropna=False):
        month_group = monthly[(monthly["Line"] == line) & (monthly["Activity / Item"] == activity)]
        peak = 0.0
        peak_month = ""
        if not month_group.empty:
            by_month = month_group.groupby(["Month ISO", "Month Label"], as_index=False)["Load"].sum()
            row = by_month.loc[by_month["Load"].idxmax()]
            peak = number(row["Load"])
            peak_month = row["Month Label"]
        req_info = parse_requirement(group["Item Requirement"].dropna().iloc[0] if group["Item Requirement"].notna().any() else "")
        disciplines = group["Discipline"].value_counts().to_dict()
        roles = sorted(set(clean(v) for v in group["Role"] if clean(v)))
        instances.append(
            {
                "line": line,
                "activity": activity,
                "testType": test_type,
                "requirement": req_info["text"],
                "pax": req_info["pax"],
                "shifts": req_info["shifts"],
                "coreFloor": req_info["coreFloor"],
                "unit": req_info["unit"],
                "roleRows": int(group["Row ID"].nunique()),
                "peakLoad": rounded(peak),
                "peakMonth": peak_month,
                "disciplines": disciplines,
                "roles": roles[:12],
            }
        )

    test_basis = []
    for test_type in sorted(set(item["testType"] for item in instances)):
        items = [item for item in instances if item["testType"] == test_type]
        core_values = [item["coreFloor"] for item in items if item["coreFloor"] is not None]
        pax_values = [item["pax"] for item in items if item["pax"] is not None]
        peak_values = [item["peakLoad"] for item in items if item["peakLoad"] is not None]
        role_values = [item["roleRows"] for item in items]
        all_disciplines = defaultdict(list)
        for item in items:
            for discipline, count in item["disciplines"].items():
                all_disciplines[discipline].append(count)
        typical_mix = {
            discipline: rounded(median(counts), 0)
            for discipline, counts in sorted(all_disciplines.items())
            if median(counts)
        }
        requirements = [item["requirement"] for item in items if item["requirement"]]
        requirement_mode = Counter(requirements).most_common(1)[0][0] if requirements else ""
        core_floor = mode_number(core_values)
        if core_floor is None:
            core_floor = rounded(median(peak_values), 0)
        test_basis.append(
            {
                "testType": test_type,
                "instances": len(items),
                "lines": sorted(set(item["line"] for item in items)),
                "basisRequirement": requirement_mode,
                "statedPax": mode_number(pax_values),
                "shiftBasis": mode_number([item["shifts"] for item in items if item["shifts"]]),
                "coreFloor": rounded(core_floor, 0) if core_floor is not None else None,
                "medianPeakLoad": rounded(median(peak_values)),
                "p75PeakLoad": rounded(percentile(peak_values, 0.75)),
                "maxPeakLoad": rounded(max(peak_values) if peak_values else 0),
                "medianRoleRows": rounded(median(role_values), 0),
                "roleMix": typical_mix,
            }
        )
    test_basis.sort(key=lambda x: (-(x["maxPeakLoad"] or 0), x["testType"]))

    matrix = []
    for test_type in sorted(set(monthly["Test Type"])):
        row = {"testType": test_type}
        for line in LINES:
            subset = monthly[(monthly["Test Type"] == test_type) & (monthly["Line"] == line)]
            if subset.empty:
                row[line] = 0
            else:
                by_month = subset.groupby("Month ISO")["Load"].sum()
                row[line] = rounded(by_month.max())
        row["totalPeak"] = rounded(sum(row[line] for line in LINES))
        matrix.append(row)
    matrix.sort(key=lambda x: -x["totalPeak"])

    demand_rows = []
    for _, row in demand_line.iterrows():
        item = {
            "month": row["Month ISO"],
            "label": row["Month Label"],
            "threshold": rounded(row["Threshold"]),
            "total": rounded(row["Total"]),
        }
        for line in LINES:
            item[line] = rounded(row[line])
        demand_rows.append(item)

    company_rows = []
    for _, row in demand_company.iterrows():
        company_rows.append(
            {
                "month": row["Month ISO"],
                "label": row["Month Label"],
                "Siemens": rounded(row["Siemens"]),
                "SC": rounded(row["SC"]),
                "TBC": rounded(row["TBC"]),
                "Unknown": rounded(row["Unknown"]),
                "total": rounded(row["Total"]),
             }
        )

    def build_demand_rows_for_view(view):
        period_col = "View Week" if view == "Week" else "View Month"
        label_col = "View Week Label" if view == "Week" else "View Month Label"
        if monthly.empty:
            return [], []
        pivot = (
            monthly.pivot_table(
                index=[period_col, label_col],
                columns="Line",
                values="Load",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
            .sort_values(period_col)
        )
        for line in LINES:
            if line not in pivot.columns:
                pivot[line] = 0.0
        pivot["Total"] = pivot[LINES].sum(axis=1)
        pivot["Threshold"] = float(len(people_roster) or len(manpower))
        rows = []
        periods = []
        for _, view_row in pivot.iterrows():
            period_iso = month_iso(view_row[period_col])
            label = clean(view_row[label_col])
            item = {
                "month": period_iso,
                "label": label,
                "threshold": rounded(view_row["Threshold"]),
                "total": rounded(view_row["Total"]),
            }
            for line in LINES:
                item[line] = rounded(view_row[line])
            rows.append(item)
            periods.append({"Month ISO": period_iso, "Month Label": label, "month": period_iso, "label": label})
        return rows, periods

    week_demand_rows, week_periods = build_demand_rows_for_view("Week")
    month_demand_rows, month_periods = build_demand_rows_for_view("Month")
    demand_rows_by_view = {"Week": week_demand_rows, "Month": month_demand_rows}
    periods_by_view = {"Week": week_periods, "Month": month_periods}

    total_peak = max((row["total"] for row in demand_rows), default=0)
    peak_month = next((row["label"] for row in demand_rows if row["total"] == total_peak), "")
    summary = {
        "source": source_label,
        "sourceMode": source_mode,
        "resourceRows": int(len(resources)),
        "monthlyRows": int(len(monthly)),
        "excludedResourceRows": excluded_resource_rows,
        "excludedMonthlyRows": excluded_monthly_rows,
        "totalManpower": int(len(people_roster) or len(manpower)),
        "tbcRows": int((resources["TBC?"] == "YES").sum()),
        "overloadPersonMonths": int(len(clash_records)),
        "personClashes": issue_summary["personClashes"],
        "tbcGaps": issue_summary["tbcGaps"],
        "fullyLoaded": issue_summary["fullyLoaded"],
        "availableResources": issue_summary["available"],
        "peakDemand": rounded(total_peak),
        "peakMonth": peak_month,
        "testTypes": len(all_test_types),
        "lines": LINES,
        "periodType": period_type,
    }

    monthly_loads_by_row = defaultdict(list)
    for _, row in monthly.iterrows():
        monthly_loads_by_row[row["Row ID"]].append(
            {
                "month": row["Month ISO"],
                "monthLabel": row["Month Label"],
                "period": row["Month ISO"],
                "periodLabel": row["Month Label"],
                "load": rounded(row["Load"]),
            }
        )

    issue_by_row = defaultdict(list)
    load_status_by_row = defaultdict(list)
    for row in issue_records:
        for row_id in row.get("rowIds", []) or []:
            if row_id:
                issue_by_row[row_id].append(row)
    for row in person_load_records:
        for row_id in row.get("rowIds", []) or []:
            if row_id:
                 load_status_by_row[row_id].append(row)

    detail_rows = resources[
        [
            "Row ID",
            "Source Row",
            "Line",
            "Test Type",
            "Activity / Item",
            "Item Requirement",
            "Role",
            "Assigned Name",
            "Company",
            "TBC?",
            "Start Label",
            "Finish Label",
            "Core Floor",
            "Requirement Unit",
        ]
    ].copy()
    detail_rows.columns = [
        "rowId",
        "sourceRow",
        "line",
        "testType",
        "activity",
        "requirement",
        "role",
        "assigned",
        "company",
        "tbc",
        "start",
        "finish",
        "coreFloor",
        "unit",
    ]
    def build_loads_by_row(view):
        period_col = "View Week" if view == "Week" else "View Month"
        label_col = "View Week Label" if view == "Week" else "View Month Label"
        loads = defaultdict(list)
        if monthly.empty:
            return loads
        grouped = (
            monthly.groupby(["Row ID", period_col, label_col], as_index=False)["Load"]
            .sum()
            .sort_values([period_col, "Row ID"])
        )
        for _, row in grouped.iterrows():
            loads[row["Row ID"]].append(
                {
                    "month": month_iso(row[period_col]),
                    "monthLabel": clean(row[label_col]),
                    "period": month_iso(row[period_col]),
                    "periodLabel": clean(row[label_col]),
                    "load": rounded(row["Load"], 2),
                }
            )
        return loads

    month_loads_by_row = build_loads_by_row("Month")
    week_loads_by_row = build_loads_by_row("Week")

    detail_records = as_records(detail_rows)
    for item in detail_records:
        period_loads = week_loads_by_row.get(item["rowId"], []) if period_type == "Week" else month_loads_by_row.get(item["rowId"], [])
        row_issues = issue_by_row.get(item["rowId"], [])
        row_loads = load_status_by_row.get(item["rowId"], [])
        item["monthlyLoads"] = month_loads_by_row.get(item["rowId"], [])
        item["monthLoads"] = month_loads_by_row.get(item["rowId"], [])
        item["weekLoads"] = week_loads_by_row.get(item["rowId"], [])
        item["periodLoads"] = period_loads
        item["roleKey"] = role_key(item.get("role", ""))
        item["clash"] = any(row.get("status") == "Clash" for row in row_issues)
        item["status"] = (
            "TBC Gap"
            if item.get("tbc") == "YES"
            else "Clash"
            if item["clash"]
            else "Fully Loaded"
            if any(row.get("status") == "Fully Loaded" for row in row_loads)
            else "Available"
        )
        item["issuePeriods"] = sorted({row.get("monthLabel", "") for row in row_issues if row.get("monthLabel")})
        item["clashMonths"] = sorted({row.get("monthLabel", "") for row in row_issues if row.get("status") == "Clash" and row.get("monthLabel")})
        item["excessLoad"] = rounded(sum(number(row.get("excessLoad", 0)) for row in row_issues), 2)
        item["suggestedFix"] = "; ".join(sorted({row.get("suggestedFix", "") for row in row_issues if row.get("suggestedFix")}))

    people_options = sorted(
        {
            clean(name)
            for name in [row["name"] for row in people_roster] + list(resources.get("Assigned Name", pd.Series(dtype=str)))
            if clean(name) and not is_tbc_name(name)
        }
    )

    return {
        "summary": summary,
        "lineColors": LINE_COLORS,
        "months": months,
        "lines": LINES,
        "manpower": manpower_summary,
        "lineMetrics": line_metrics,
        "demandRows": demand_rows,
        "demandRowsByView": demand_rows_by_view,
        "periodsByView": periods_by_view,
        "companyRows": company_rows,
        "testBasis": CORE_TEAM_BASIS,
        "testLineMatrix": matrix,
        "testTypes": all_test_types,
        "instances": instances,
        "detailRows": detail_records,
        "personLoads": person_load_records,
        "clashRows": clash_records,
        "issueRows": issue_records,
        "issueSummary": issue_summary,
        "peopleRoster": people_roster,
        "peopleOptions": people_options,
        "periodType": period_type,
        "viewMode": period_type,
        "viewModes": ["Week", "Month"],
        "weeklyExact": period_type == "Week",
    }


def set_sheet_layout(ws, widths):
    ws.freeze_panes = "A2"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_table(ws, headers, rows, widths=None, red_flag_col=None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    clash_fill = PatternFill("solid", fgColor="FCE4E4")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        is_red = bool(red_flag_col and row.get(red_flag_col))
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_red:
                cell.fill = clash_fill

    if widths:
        set_sheet_layout(ws, widths)
    ws.auto_filter.ref = ws.dimensions


def export_results_workbook(payload):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    summary = payload.get("summary", {})
    issue_summary = payload.get("issueSummary", {})
    period_name = summary.get("periodType") or payload.get("periodType") or "Month"
    summary_rows = [
        {"Metric": "Source", "Value": summary.get("source", "")},
        {"Metric": "Source Mode", "Value": summary.get("sourceMode", "")},
        {"Metric": "Counted Resource Rows", "Value": summary.get("resourceRows", 0)},
        {"Metric": "Summary/Header Rows Excluded", "Value": summary.get("excludedResourceRows", 0)},
        {"Metric": f"{period_name} Records", "Value": summary.get("monthlyRows", 0)},
        {"Metric": "Total Manpower Master", "Value": summary.get("totalManpower", 0)},
        {"Metric": "TBC Gaps", "Value": issue_summary.get("tbcGaps", summary.get("tbcGaps", 0))},
        {"Metric": "Person Clashes", "Value": issue_summary.get("personClashes", summary.get("personClashes", 0))},
        {"Metric": "Fully Loaded Resources", "Value": issue_summary.get("fullyLoaded", summary.get("fullyLoaded", 0))},
        {"Metric": "Available Resources", "Value": issue_summary.get("available", summary.get("availableResources", 0))},
        {"Metric": "Peak Demand", "Value": summary.get("peakDemand", 0)},
        {"Metric": f"Peak {period_name}", "Value": summary.get("peakMonth", "")},
    ]
    write_table(ws, ["Metric", "Value"], summary_rows, [30, 40])

    ws = wb.create_sheet("Line Summary")
    line_rows = []
    for row in payload.get("lineMetrics", []):
        line_rows.append(
            {
                "Line": row.get("line", ""),
                "Peak Demand": row.get("peakDemand", 0),
                "Peak Month": row.get("peakMonth", ""),
                "Available": row.get("available", 0),
                "Gap": row.get("gap", 0),
                "TBC Rows": row.get("tbcRows", 0),
            }
        )
    write_table(ws, ["Line", "Peak Demand", "Peak Month", "Available", "Gap", "TBC Rows"], line_rows, [12, 15, 14, 14, 12, 12])

    ws = wb.create_sheet("Demand by Line")
    demand_headers = [period_name, "DTL", "JRL", "CRL", "RTS", "Total", "Threshold"]
    demand_rows = []
    for row in payload.get("demandRows", []):
        demand_rows.append(
            {
                period_name: row.get("label", ""),
                "DTL": row.get("DTL", 0),
                "JRL": row.get("JRL", 0),
                "CRL": row.get("CRL", 0),
                "RTS": row.get("RTS", 0),
                "Total": row.get("total", 0),
                "Threshold": row.get("threshold", 0),
            }
        )
    write_table(ws, demand_headers, demand_rows, [12, 10, 10, 10, 10, 10, 12])

    ws = wb.create_sheet("Core Team Standard")
    core_headers = ["Test Type", "Roles Required", "Siemens Pax", "Subcon Pax", "Total Pax"]
    core_rows = []
    for row in payload.get("testBasis", CORE_TEAM_BASIS):
        core_rows.append(
            {
                "Test Type": row.get("testType", ""),
                "Roles Required": row.get("rolesRequired", ""),
                "Siemens Pax": row.get("siemensPax", 0),
                "Subcon Pax": row.get("subconPax", 0),
                "Total Pax": row.get("totalPax", 0),
            }
        )
    write_table(ws, core_headers, core_rows, [34, 90, 14, 14, 12])

    ws = wb.create_sheet("Assignment Detail")
    assignment_headers = [
        "Line",
        "Test Type",
        "Activity",
        "Role",
        "Assigned Name",
        "Company",
        "TBC",
        "Status",
        "Issue Periods",
        "Excess Load",
        "Suggested Fix",
        "Start",
        "Finish",
        "Requirement",
    ]
    assignment_rows = []
    for row in payload.get("detailRows", []):
        assignment_rows.append(
            {
                "Line": row.get("line", ""),
                "Test Type": row.get("testType", ""),
                "Activity": row.get("activity", ""),
                "Role": row.get("role", ""),
                "Assigned Name": row.get("assigned", ""),
                "Company": row.get("company", ""),
                "TBC": row.get("tbc", ""),
                "Status": row.get("status", "Clash" if row.get("clash") else "Available"),
                "Issue Periods": ", ".join(row.get("issuePeriods", []) or row.get("clashMonths", []) or []),
                "Excess Load": row.get("excessLoad", 0),
                "Suggested Fix": row.get("suggestedFix", ""),
                "Start": row.get("start", ""),
                "Finish": row.get("finish", ""),
                "Requirement": row.get("requirement", ""),
                "_red": row.get("status") in {"Clash", "TBC Gap"} or bool(row.get("clash")),
            }
        )
    write_table(ws, assignment_headers, assignment_rows, [10, 26, 36, 28, 22, 14, 10, 14, 24, 12, 36, 12, 12, 26], red_flag_col="_red")

    ws = wb.create_sheet("Issue List")
    issue_headers = ["Person", "Company", period_name, "Role", "Line", "Test Type", "Load", "Max Weekly Load", "Status", "Excess Load", "Excess Days", "Suggested Fix"]
    issue_rows = []
    for row in payload.get("issueRows", payload.get("clashRows", [])):
        issue_rows.append(
            {
                "Person": row.get("assigned", ""),
                "Company": row.get("company", ""),
                period_name: row.get("weekLabel") or row.get("monthLabel", ""),
                "Role": row.get("role", ""),
                "Line": row.get("line", ""),
                "Test Type": row.get("testType", ""),
                "Load": row.get("load", 0),
                "Max Weekly Load": row.get("threshold", 1),
                "Status": row.get("status", "Clash" if row.get("clash") else ""),
                "Excess Load": row.get("excessLoad", 0),
                "Excess Days": row.get("excessDays", 0),
                "Suggested Fix": row.get("suggestedFix", ""),
                "_red": row.get("status") in {"Clash", "TBC Gap"} or bool(row.get("clash")),
            }
        )
    write_table(ws, issue_headers, issue_rows, [24, 14, 12, 28, 10, 26, 10, 16, 14, 12, 12, 44], red_flag_col="_red")

    ws = wb.create_sheet("Person Load")
    load_headers = ["Person", "Company", period_name, "Role", "Load", "Max Weekly Load", "Status", "Excess Load", "Suggested Fix"]
    load_rows = []
    for row in payload.get("personLoads", []):
        load_rows.append(
            {
                "Person": row.get("assigned", ""),
                "Company": row.get("company", ""),
                period_name: row.get("weekLabel") or row.get("monthLabel", ""),
                "Role": row.get("role", ""),
                "Load": row.get("load", 0),
                "Max Weekly Load": row.get("threshold", 1),
                "Status": row.get("status", "Clash" if row.get("clash") else ""),
                "Excess Load": row.get("excessLoad", 0),
                "Suggested Fix": row.get("suggestedFix", ""),
                "_red": row.get("status") in {"Clash", "TBC Gap"} or bool(row.get("clash")),
            }
        )
    write_table(ws, load_headers, load_rows, [24, 14, 12, 28, 10, 16, 14, 12, 42], red_flag_col="_red")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def json_safe(value):
    if value is None:
        return None
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(v) for v in value]
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, (str, int, bool)):
        return value
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return str(value)


def write_static_dashboard(payload, output_path):
    data_json = json.dumps(json_safe(payload), ensure_ascii=False).replace("</", "<\\/")
    html = STATIC_DASHBOARD_HTML.replace("__DASHBOARD_DATA__", data_json)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    return output_path


def generate_outputs(workbook_path, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = build_data(workbook_path=workbook_path, source_label=Path(workbook_path).name)
    html_path = write_static_dashboard(payload, output_dir / "tc_resource_allocation_dashboard.html")
    report_path = output_dir / "manpower_allocation_report.xlsx"
    report_path.write_bytes(export_results_workbook(payload))
    code_path = output_dir / "resource_dashboard.py"
    if Path(__file__).resolve() != code_path.resolve():
        shutil.copy2(Path(__file__), code_path)
    return {
        "payload": payload,
        "html": html_path,
        "report": report_path,
        "code": code_path,
    }


STATIC_DASHBOARD_HTML = r"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>T&C Resource Allocation Dashboard</title>
  <style>
    :root {
      --bg: #f6f8fb;
      --panel: #ffffff;
      --ink: #23303d;
      --muted: #637083;
      --line: #d8e0e8;
      --accent: #0f766e;
      --accent-soft: #e7f4f2;
      --danger: #c2413b;
      --danger-soft: #fdebea;
      --warn: #b7791f;
      --warn-soft: #fff4db;
      --ok: #237a4b;
      --ok-soft: #e8f6ee;
      --full: #2666a3;
      --full-soft: #e8f1fa;
      --shadow: 0 16px 36px rgba(33, 45, 58, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font: 14px/1.45 "Segoe UI", Arial, sans-serif;
    }
    header {
      padding: 18px 24px 14px;
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      position: sticky;
      top: 0;
      z-index: 20;
    }
    .title-row {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 16px;
      flex-wrap: wrap;
    }
    h1 {
      margin: 0;
      font-size: 24px;
      letter-spacing: 0;
    }
    .subtitle {
      margin-top: 4px;
      color: var(--muted);
    }
    main {
      padding: 18px 24px 32px;
      max-width: 1480px;
      margin: 0 auto;
    }
    .controls {
      display: grid;
      grid-template-columns: repeat(5, minmax(150px, 1fr));
      gap: 10px;
      margin-top: 14px;
    }
    label {
      display: grid;
      gap: 5px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      text-transform: uppercase;
    }
    select, input, button {
      font: inherit;
      border: 1px solid var(--line);
      border-radius: 7px;
      background: #fff;
      color: var(--ink);
      min-height: 36px;
    }
    select, input { padding: 7px 9px; width: 100%; }
    button {
      padding: 7px 11px;
      cursor: pointer;
      font-weight: 600;
    }
    button.primary {
      background: var(--accent);
      color: white;
      border-color: var(--accent);
    }
    button.ghost {
      background: #fff;
    }
    .button-row {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: end;
    }
    .grid {
      display: grid;
      gap: 14px;
    }
    .kpis {
      grid-template-columns: repeat(6, minmax(130px, 1fr));
      margin-bottom: 14px;
    }
    .card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 14px;
    }
    .kpi .label {
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      text-transform: uppercase;
    }
    .kpi .value {
      margin-top: 6px;
      font-size: 28px;
      font-weight: 750;
    }
    .kpi .note {
      margin-top: 4px;
      color: var(--muted);
      min-height: 20px;
    }
    .two-col {
      grid-template-columns: minmax(0, 1.7fr) minmax(320px, 0.8fr);
      align-items: start;
    }
    h2 {
      margin: 0 0 12px;
      font-size: 17px;
    }
    .chart-list {
      display: grid;
      gap: 8px;
      max-height: 420px;
      overflow: auto;
      padding-right: 4px;
    }
    .bar-row {
      display: grid;
      grid-template-columns: 72px minmax(120px, 1fr) 58px;
      gap: 8px;
      align-items: center;
    }
    .bar-shell {
      height: 20px;
      background: #edf1f5;
      border-radius: 6px;
      overflow: hidden;
      display: flex;
    }
    .bar-total {
      display: flex;
      height: 100%;
      border-radius: 6px;
      overflow: hidden;
    }
    .seg { height: 100%; min-width: 2px; }
    .small {
      color: var(--muted);
      font-size: 12px;
    }
    .legend {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-bottom: 10px;
    }
    .legend span {
      display: inline-flex;
      gap: 6px;
      align-items: center;
      color: var(--muted);
      font-size: 12px;
    }
    .swatch {
      width: 12px;
      height: 12px;
      border-radius: 3px;
      display: inline-block;
    }
    table {
      width: 100%;
      border-collapse: collapse;
    }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 8px 9px;
      text-align: left;
      vertical-align: top;
    }
    th {
      font-size: 12px;
      color: var(--muted);
      text-transform: uppercase;
      background: #f9fbfd;
      position: sticky;
      top: 0;
      z-index: 1;
    }
    .table-wrap {
      max-height: 560px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .status {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 3px 8px;
      font-size: 12px;
      font-weight: 700;
      white-space: nowrap;
    }
    .status.Clash { color: var(--danger); background: var(--danger-soft); }
    .status.TBCGap { color: var(--warn); background: var(--warn-soft); }
    .status.FullyLoaded { color: var(--full); background: var(--full-soft); }
    .status.Available { color: var(--ok); background: var(--ok-soft); }
    .tabs {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 16px 0 10px;
    }
    .tab-btn {
      border-color: var(--line);
      background: #fff;
    }
    .tab-btn.active {
      background: var(--accent-soft);
      color: var(--accent);
      border-color: #9ccfca;
    }
    .panel { display: none; }
    .panel.active { display: block; }
    .muted-box {
      padding: 9px 11px;
      border-radius: 7px;
      background: #f9fbfd;
      border: 1px solid var(--line);
      color: var(--muted);
      margin-bottom: 10px;
    }
    .assignment-select {
      min-width: 150px;
    }
    .capacity-input {
      width: 76px;
    }
    .fix-button {
      white-space: nowrap;
      min-height: 30px;
      padding: 5px 8px;
    }
    .nowrap { white-space: nowrap; }
    .right { text-align: right; }
    @media (max-width: 1050px) {
      .controls, .kpis, .two-col { grid-template-columns: 1fr 1fr; }
    }
    @media (max-width: 700px) {
      header, main { padding-left: 14px; padding-right: 14px; }
      .controls, .kpis, .two-col { grid-template-columns: 1fr; }
      .bar-row { grid-template-columns: 64px 1fr 48px; }
      h1 { font-size: 21px; }
    }
  </style>
</head>
<body>
  <header>
    <div class="title-row">
      <div>
        <h1>T&C Resource Allocation Dashboard</h1>
        <div class="subtitle" id="sourceMeta"></div>
      </div>
      <div class="button-row">
        <button class="primary" id="saveBtn">Save Edits</button>
        <button class="ghost" id="resetBtn">Reset</button>
        <button class="ghost" id="issueCsvBtn">Download Issues CSV</button>
        <button class="ghost" id="allocationCsvBtn">Download Edited Allocation CSV</button>
      </div>
    </div>
    <div class="controls">
      <label>Line
        <select id="lineFilter"></select>
      </label>
      <label>Status
        <select id="statusFilter">
          <option>All</option>
          <option>Clash</option>
          <option>TBC Gap</option>
          <option>Fully Loaded</option>
          <option>Available</option>
        </select>
      </label>
      <label>View
        <select id="viewMode">
          <option>Month</option>
          <option>Week</option>
        </select>
      </label>
      <label>Search
        <input id="searchBox" type="search" placeholder="Person, role, activity">
      </label>
      <label>Displayed Rows
        <select id="rowLimit">
          <option value="120">120</option>
          <option value="250" selected>250</option>
          <option value="600">All</option>
        </select>
      </label>
    </div>
  </header>

  <main>
    <section class="grid kpis" id="kpis"></section>

    <section class="grid two-col">
      <div class="card">
        <h2>Demand by Line</h2>
        <div class="legend" id="legend"></div>
        <div class="chart-list" id="demandChart"></div>
      </div>
      <div class="card">
        <h2>Management Notes</h2>
        <div id="managementNotes"></div>
      </div>
    </section>

    <nav class="tabs" aria-label="Dashboard sections">
      <button class="tab-btn active" data-tab="issues">Clash Detection</button>
      <button class="tab-btn" data-tab="editor">Allocation Editor</button>
      <button class="tab-btn" data-tab="people">Capacity Editor</button>
      <button class="tab-btn" data-tab="lines">Line Summary</button>
      <button class="tab-btn" data-tab="matrix">Test Matrix</button>
    </nav>

    <section class="panel active card" id="panel-issues">
      <h2>Clash Detection and Suggested Fixes</h2>
      <div class="muted-box" id="issueNote"></div>
      <div class="table-wrap"><table id="issuesTable"></table></div>
    </section>

    <section class="panel card" id="panel-editor">
      <h2>Editable Allocation Detail</h2>
      <div class="muted-box" id="editorNote"></div>
      <div class="table-wrap"><table id="editorTable"></table></div>
    </section>

    <section class="panel card" id="panel-people">
      <h2>Editable Capacity</h2>
      <div class="muted-box">Adjust max load to test adequacy. These edits update the clash calculation immediately in this browser.</div>
      <div class="table-wrap"><table id="peopleTable"></table></div>
    </section>

    <section class="panel card" id="panel-lines">
      <h2>Line Summary</h2>
      <div class="table-wrap"><table id="lineTable"></table></div>
    </section>

    <section class="panel card" id="panel-matrix">
      <h2>Test Type Matrix</h2>
      <div class="table-wrap"><table id="matrixTable"></table></div>
    </section>
  </main>

  <script>
    const INITIAL_DATA = __DASHBOARD_DATA__;
    const STORAGE_KEY = "tc-resource-dashboard-edits-v1";
    const clone = (value) => JSON.parse(JSON.stringify(value));
    const state = {
      data: INITIAL_DATA,
      rows: clone(INITIAL_DATA.detailRows || []),
      people: clone(INITIAL_DATA.peopleRoster || []),
      line: "All",
      status: "All",
      view: INITIAL_DATA.periodType || "Month",
      search: "",
      limit: 250,
      activeTab: "issues"
    };
    let derived = null;

    const lineColors = Object.assign({ DTL: "#2563eb", JRL: "#0f766e", CRL: "#d6a400", RTS: "#303b46" }, INITIAL_DATA.lineColors || {});

    function esc(value) {
      return String(value ?? "").replace(/[&<>"']/g, (char) => ({
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        '"': "&quot;",
        "'": "&#39;"
      }[char]));
    }
    function clean(value) {
      return String(value ?? "").replace(/\s+/g, " ").trim();
    }
    function num(value, fallback = 0) {
      const n = Number(value);
      return Number.isFinite(n) ? n : fallback;
    }
    function fmt(value, digits = 1) {
      const n = num(value);
      if (Math.abs(n - Math.round(n)) < 1e-9) return String(Math.round(n));
      return n.toFixed(digits).replace(/\.0+$/, "");
    }
    function isTbcName(value) {
      const text = clean(value).toUpperCase();
      return !text || ["TBC", "TBD", "NA", "N/A"].includes(text) || text.includes("TBC");
    }
    function addUnique(list, item) {
      if (item && !list.includes(item)) list.push(item);
    }
    function roleKeys(value) {
      const text = clean(value).toUpperCase();
      const keys = [];
      if (!text) return keys;
      if (text.includes("SUBCON") || /\bSC\b/.test(text)) addUnique(keys, "SUBCON");
      if (text.includes("COMMS") || text.includes("NETWORK") || text.includes("RCS")) addUnique(keys, "COMMS T&C ENGINEER");
      if (text.includes("TRAIN")) addUnique(keys, "TRAIN T&C ENGINEER");
      if (text.includes("CSF")) addUnique(keys, "CSF T&C ENGINEER");
      if (text.includes("SIG") || text.includes("IXL")) addUnique(keys, "SIG T&C ENGINEER");
      if (text.includes("ATS")) addUnique(keys, "ATS T&C ENGINEER");
      if (text.includes("ATC")) addUnique(keys, "ATC T&C ENGINEER");
      if (text.includes("MANAGER")) addUnique(keys, "MANAGER");
      if (text.includes("COORDINATOR")) addUnique(keys, "COORDINATOR");
      if (text.includes("DESIGN")) addUnique(keys, "DESIGN ENGINEER");
      if (!keys.length) addUnique(keys, text.replace(/\b\d+\b/g, "").replace(/\s+/g, " ").trim());
      return keys.sort();
    }
    function preferredCompany(role) {
      return roleKeys(role).includes("SUBCON") ? "SC" : "Siemens";
    }
    function statusClass(status) {
      return clean(status).replace(/\s+/g, "");
    }
    function statusBadge(status) {
      return `<span class="status ${statusClass(status)}">${esc(status)}</span>`;
    }
    function personKey(name) {
      return clean(name).toUpperCase();
    }
    function peopleByName() {
      const map = new Map();
      for (const person of state.people) map.set(personKey(person.name), person);
      return map;
    }
    function personOptions(selected) {
      const selectedText = clean(selected) || "TBC";
      const names = ["TBC", ...state.people.map((person) => person.name).sort((a, b) => a.localeCompare(b))];
      return names.map((name) => `<option value="${esc(name)}"${clean(name) === selectedText ? " selected" : ""}>${esc(name)}</option>`).join("");
    }
    function rowLoads(row) {
      const source = state.view === "Week" ? (row.weekLoads || row.periodLoads || []) : (row.monthLoads || row.periodLoads || []);
      return source.length ? source : (row.periodLoads || []);
    }
    function rowPeakLoad(row, period) {
      let total = 0;
      for (const load of rowLoads(row)) {
        if (!period || load.period === period || load.month === period) total += num(load.load);
      }
      return total;
    }
    function markRow(rowStates, rowId, status) {
      const rank = { "Available": 0, "Fully Loaded": 1, "TBC Gap": 2, "Clash": 3 };
      const current = rowStates.get(rowId) || "Available";
      if ((rank[status] ?? 0) > (rank[current] ?? 0)) rowStates.set(rowId, status);
    }
    function makeSuggestion(issue, loadIndex) {
      const targetKeys = roleKeys(issue.role);
      const targetLabel = targetKeys.length ? targetKeys.join("/") : "matching role";
      const preferred = preferredCompany(issue.role);
      const needed = issue.status === "Clash" ? num(issue.excessLoad) : num(issue.load);
      const candidates = [];
      for (const person of state.people) {
        if (personKey(person.name) === personKey(issue.assigned)) continue;
        const pKeys = (person.roleKeys && person.roleKeys.length) ? person.roleKeys : roleKeys(person.role || person.roleKey || "");
        if (targetKeys.length && !pKeys.some((key) => targetKeys.includes(key))) continue;
        const currentLoad = loadIndex.get(`${personKey(person.name)}|${issue.period}`) || 0;
        const maxLoad = num(person.maxWeeklyLoad, 1);
        const capacity = maxLoad - currentLoad;
        if (capacity <= 1e-9) continue;
        const lineMatch = issue.line && (person.lines || []).includes(issue.line) ? 1 : 0;
        const companyMatch = preferred && clean(person.company) === preferred ? 1 : 0;
        const matched = pKeys.filter((key) => targetKeys.includes(key));
        candidates.push({ person, capacity, lineMatch, companyMatch, matched });
      }
      candidates.sort((a, b) => b.companyMatch - a.companyMatch || b.lineMatch - a.lineMatch || b.capacity - a.capacity || a.person.name.localeCompare(b.person.name));
      if (candidates.length) {
        const best = candidates[0];
        const roleText = best.matched.length ? best.matched.join("/") : clean(best.person.roleKey || best.person.role || "");
        const qualifier = [roleText, clean(best.person.line || "")].filter(Boolean).join(", ");
        const suffix = qualifier ? ` (${qualifier})` : "";
        if (best.capacity + 1e-9 >= needed) {
          const action = issue.status === "TBC Gap" ? "Assign" : "Move";
          const canApply = canApplyCandidate(issue, best.person.name, loadIndex);
          return {
            text: `${action} ${fmt(needed, 2)} load to ${best.person.name}${suffix}; ${fmt(best.capacity, 2)} capacity available in ${issue.label}.${canApply ? "" : " Check other loaded periods before editing this row."}`,
            candidate: canApply ? best.person.name : "",
            needed
          };
        }
        return {
          text: `Split: ${best.person.name}${suffix} can take ${fmt(best.capacity, 2)} in ${issue.label}; add another ${targetLabel} resource or shift ${fmt(Math.max(needed - best.capacity, 0), 2)} load.`,
          candidate: "",
          needed
        };
      }
      const action = issue.status === "Clash" ? `add ${fmt(needed, 2)} load or move one activity to another period` : `assign a new person or shift ${fmt(needed, 2)} load`;
      return { text: `No spare ${targetLabel} capacity found in ${issue.label}; ${action}.`, candidate: "", needed };
    }
    function rowsForIssueApply(issue) {
      let remaining = num(issue.needed || issue.excessLoad || issue.load);
      const rows = state.rows
        .filter((row) => (issue.rowIds || []).includes(row.rowId))
        .sort((a, b) => rowPeakLoad(b, issue.period) - rowPeakLoad(a, issue.period));
      const selected = [];
      for (const row of rows) {
        if (remaining <= 1e-9) break;
        selected.push(row);
        remaining -= Math.max(rowPeakLoad(row, issue.period), 0.1);
      }
      return selected;
    }
    function canApplyCandidate(issue, candidate, loadIndex) {
      if (!candidate) return false;
      const person = peopleByName().get(personKey(candidate));
      if (!person) return false;
      const maxLoad = num(person.maxWeeklyLoad, 1);
      const selectedRows = rowsForIssueApply(issue);
      if (!selectedRows.length) return false;
      for (const row of selectedRows) {
        for (const load of rowLoads(row)) {
          const amount = num(load.load);
          if (amount <= 0) continue;
          const period = load.period || load.month || "";
          const current = loadIndex.get(`${personKey(candidate)}|${period}`) || 0;
          if (current + amount > maxLoad + 1e-9) return false;
        }
      }
      return true;
    }
    function computeDerived() {
      const byName = peopleByName();
      const groups = new Map();
      const loadIndex = new Map();
      const rowStates = new Map();
      const issues = [];
      for (const row of state.rows) {
        const assigned = clean(row.assigned) || "TBC";
        const person = byName.get(personKey(assigned));
        const company = isTbcName(assigned) ? "TBC" : clean(person?.company || row.company || "Unknown");
        row.company = company;
        for (const load of rowLoads(row)) {
          const amount = num(load.load);
          if (amount <= 0) continue;
          const period = load.period || load.month || "";
          const label = load.periodLabel || load.monthLabel || period;
          if (isTbcName(assigned)) {
            const issue = {
              status: "TBC Gap",
              assigned: "TBC",
              company: "TBC",
              period,
              label,
              month: period,
              monthLabel: label,
              role: clean(row.role),
              line: clean(row.line),
              testType: clean(row.testType),
              activity: clean(row.activity),
              load: amount,
              threshold: 1,
              excessLoad: amount,
              rowIds: [row.rowId],
              rowId: row.rowId
            };
            issues.push(issue);
            markRow(rowStates, row.rowId, "TBC Gap");
            continue;
          }
          const key = `${personKey(assigned)}|${period}`;
          if (!groups.has(key)) {
            groups.set(key, {
              assigned,
              company,
              period,
              label,
              load: 0,
              roles: new Set(),
              lines: new Set(),
              rowIds: new Set(),
              rows: []
            });
          }
          const group = groups.get(key);
          group.load += amount;
          group.roles.add(clean(row.role));
          group.lines.add(clean(row.line));
          group.rowIds.add(row.rowId);
          group.rows.push(row);
        }
      }
      const personLoads = [];
      for (const group of groups.values()) {
        const person = byName.get(personKey(group.assigned));
        const threshold = num(person?.maxWeeklyLoad, 1);
        const status = group.load > threshold + 1e-9 ? "Clash" : (Math.abs(group.load - threshold) <= 1e-9 ? "Fully Loaded" : "Available");
        const lineList = [...group.lines].filter(Boolean).sort();
        const record = {
          status,
          assigned: group.assigned,
          company: group.company,
          period: group.period,
          label: group.label,
          month: group.period,
          monthLabel: group.label,
          role: [...group.roles].filter(Boolean).sort().slice(0, 4).join(", "),
          line: lineList.length === 1 ? lineList[0] : "",
          lines: lineList.join(", "),
          load: group.load,
          threshold,
          excessLoad: Math.max(group.load - threshold, 0),
          rowIds: [...group.rowIds]
        };
        personLoads.push(record);
        loadIndex.set(`${personKey(group.assigned)}|${group.period}`, group.load);
        for (const row of group.rows) markRow(rowStates, row.rowId, status);
        if (status === "Clash") issues.push(record);
      }
      for (const row of state.rows) {
        if (!rowStates.has(row.rowId)) rowStates.set(row.rowId, "Available");
      }
      for (const issue of issues) {
        const suggestion = makeSuggestion(issue, loadIndex);
        issue.suggestedFix = suggestion.text;
        issue.candidate = suggestion.candidate;
        issue.needed = suggestion.needed;
      }
      issues.sort((a, b) => {
        const statusRank = { "Clash": 0, "TBC Gap": 1 };
        return (a.period || "").localeCompare(b.period || "") ||
          (statusRank[a.status] ?? 9) - (statusRank[b.status] ?? 9) ||
          num(b.excessLoad) - num(a.excessLoad) ||
          clean(a.assigned).localeCompare(clean(b.assigned));
      });
      personLoads.sort((a, b) => (a.period || "").localeCompare(b.period || "") || clean(a.assigned).localeCompare(clean(b.assigned)));
      return { issues, personLoads, rowStates, loadIndex };
    }
    function filteredRows() {
      const query = state.search.toLowerCase();
      return state.rows.filter((row) => {
        const status = derived.rowStates.get(row.rowId) || "Available";
        if (state.line !== "All" && row.line !== state.line) return false;
        if (state.status !== "All" && status !== state.status) return false;
        if (!query) return true;
        return [row.assigned, row.role, row.activity, row.testType, row.line].join(" ").toLowerCase().includes(query);
      });
    }
    function filteredIssues() {
      const query = state.search.toLowerCase();
      return derived.issues.filter((issue) => {
        if (state.line !== "All" && issue.line && issue.line !== state.line) return false;
        if (state.status !== "All" && issue.status !== state.status) return false;
        if (!query) return true;
        return [issue.assigned, issue.role, issue.activity, issue.testType, issue.line, issue.suggestedFix].join(" ").toLowerCase().includes(query);
      });
    }
    function renderKpis() {
      const summary = state.data.summary || {};
      const namedClashes = derived.issues.filter((x) => x.status === "Clash").length;
      const tbcGaps = derived.issues.filter((x) => x.status === "TBC Gap").length;
      const fullyLoaded = derived.personLoads.filter((x) => x.status === "Fully Loaded").length;
      const available = derived.personLoads.filter((x) => x.status === "Available").length;
      const adequacy = namedClashes ? "Action needed" : (tbcGaps ? "TBC open" : "Adequate");
      const cards = [
        ["Peak Demand", fmt(summary.peakDemand), summary.peakMonth || ""],
        ["Named Clashes", fmt(namedClashes, 0), "Over threshold"],
        ["TBC Gaps", fmt(tbcGaps, 0), "Unassigned load rows"],
        ["Fully Loaded", fmt(fullyLoaded, 0), "At max load"],
        ["Available", fmt(available, 0), "Below max load"],
        ["Adequacy", adequacy, `${fmt(summary.totalManpower || state.people.length, 0)} people in roster`]
      ];
      document.getElementById("kpis").innerHTML = cards.map(([label, value, note]) => `
        <div class="card kpi">
          <div class="label">${esc(label)}</div>
          <div class="value">${esc(value)}</div>
          <div class="note">${esc(note)}</div>
        </div>`).join("");
    }
    function renderDemandChart() {
      const rows = (state.data.demandRowsByView || {})[state.view] || state.data.demandRows || [];
      const lines = state.line === "All" ? (state.data.lines || []) : [state.line];
      const maxTotal = Math.max(1, ...rows.map((row) => lines.reduce((sum, line) => sum + num(row[line]), 0)));
      document.getElementById("legend").innerHTML = lines.map((line) => `<span><i class="swatch" style="background:${lineColors[line] || "#999"}"></i>${esc(line)}</span>`).join("");
      document.getElementById("demandChart").innerHTML = rows.map((row) => {
        const total = lines.reduce((sum, line) => sum + num(row[line]), 0);
        const totalWidth = Math.max(1, Math.min(100, (total / maxTotal) * 100));
        const segs = lines.map((line) => {
          const value = num(row[line]);
          const width = total ? (value / total) * 100 : 0;
          return `<div class="seg" title="${esc(line)} ${fmt(value)}" style="width:${width}%; background:${lineColors[line] || "#999"}"></div>`;
        }).join("");
        return `<div class="bar-row">
          <div class="small nowrap">${esc(row.label)}</div>
          <div class="bar-shell"><div class="bar-total" style="width:${totalWidth}%">${segs}</div></div>
          <div class="small right">${fmt(total)}</div>
        </div>`;
      }).join("");
    }
    function renderNotes() {
      const metrics = state.data.lineMetrics || [];
      const summary = state.data.summary || {};
      const worst = [...metrics].sort((a, b) => num(b.gap) - num(a.gap))[0];
      const topIssue = derived.issues[0];
      const notes = [
        `<strong>Source:</strong> ${esc(summary.source || "")} (${esc(summary.sourceMode || "")})`,
        `<strong>Rows counted:</strong> ${fmt(summary.resourceRows || 0, 0)} allocation rows and ${fmt(summary.monthlyRows || 0, 0)} period load records.`,
        worst ? `<strong>Largest line gap:</strong> ${esc(worst.line)} peaks at ${fmt(worst.peakDemand)} vs ${fmt(worst.available)} available (${esc(worst.peakMonth)}).` : "",
        topIssue ? `<strong>First action:</strong> ${esc(topIssue.status)} for ${esc(topIssue.assigned)} in ${esc(topIssue.label)}. ${esc(topIssue.suggestedFix)}` : "<strong>Status:</strong> No open issues after current edits."
      ].filter(Boolean);
      document.getElementById("managementNotes").innerHTML = notes.map((note) => `<p>${note}</p>`).join("");
    }
    function renderIssues() {
      const rows = filteredIssues();
      document.getElementById("issueNote").textContent = `${rows.length} issue rows after filters. Use Apply to test the suggested reassignment in the editable allocation table.`;
      const limited = rows.slice(0, state.limit);
      document.getElementById("issuesTable").innerHTML = `
        <thead><tr>
          <th>Status</th><th>Period</th><th>Person</th><th>Line</th><th>Role</th><th class="right">Load</th><th class="right">Max</th><th>Suggested Fix</th><th></th>
        </tr></thead>
        <tbody>${limited.map((issue, index) => `
          <tr>
            <td>${statusBadge(issue.status)}</td>
            <td class="nowrap">${esc(issue.label)}</td>
            <td>${esc(issue.assigned)}</td>
            <td>${esc(issue.line || issue.lines || "")}</td>
            <td>${esc(issue.role)}</td>
            <td class="right">${fmt(issue.load, 2)}</td>
            <td class="right">${fmt(issue.threshold, 2)}</td>
            <td>${esc(issue.suggestedFix)}</td>
            <td>${issue.candidate ? `<button class="fix-button" data-apply-issue="${index}">Apply</button>` : ""}</td>
          </tr>`).join("")}</tbody>`;
      document.querySelectorAll("[data-apply-issue]").forEach((button) => {
        button.addEventListener("click", () => applyIssue(limited[num(button.dataset.applyIssue)]));
      });
    }
    function renderEditor() {
      const rows = filteredRows();
      const limited = rows.slice(0, state.limit);
      document.getElementById("editorNote").textContent = `Showing ${limited.length} of ${rows.length} filtered allocation rows. Assignment edits recalculate all clash and TBC statuses immediately.`;
      document.getElementById("editorTable").innerHTML = `
        <thead><tr>
          <th>Status</th><th>Line</th><th>Test Type</th><th>Activity</th><th>Role</th><th>Assigned</th><th>Periods</th><th class="right">Peak Load</th>
        </tr></thead>
        <tbody>${limited.map((row) => {
          const status = derived.rowStates.get(row.rowId) || "Available";
          const periods = rowLoads(row).map((load) => `${load.periodLabel || load.monthLabel}: ${fmt(load.load, 2)}`).join(", ");
          return `<tr>
            <td>${statusBadge(status)}</td>
            <td>${esc(row.line)}</td>
            <td>${esc(row.testType)}</td>
            <td>${esc(row.activity)}</td>
            <td>${esc(row.role)}</td>
            <td><select class="assignment-select" data-assign-row="${esc(row.rowId)}">${personOptions(row.assigned)}</select></td>
            <td class="small">${esc(periods)}</td>
            <td class="right">${fmt(Math.max(0, ...rowLoads(row).map((load) => num(load.load))), 2)}</td>
          </tr>`;
        }).join("")}</tbody>`;
      document.querySelectorAll("[data-assign-row]").forEach((select) => {
        select.addEventListener("change", () => {
          const row = state.rows.find((item) => item.rowId === select.dataset.assignRow);
          if (!row) return;
          row.assigned = select.value;
          const person = peopleByName().get(personKey(select.value));
          row.company = isTbcName(select.value) ? "TBC" : clean(person?.company || "Unknown");
          renderAll();
        });
      });
    }
    function renderPeople() {
      document.getElementById("peopleTable").innerHTML = `
        <thead><tr><th>Name</th><th>Company</th><th>Line</th><th>Role</th><th>Role Keys</th><th class="right">Max Load</th></tr></thead>
        <tbody>${state.people.map((person) => `
          <tr>
            <td>${esc(person.name)}</td>
            <td>${esc(person.company)}</td>
            <td>${esc(person.line || "")}</td>
            <td>${esc(person.role || "")}</td>
            <td>${esc((person.roleKeys || []).join(", "))}</td>
            <td class="right"><input class="capacity-input" type="number" min="0" step="0.1" value="${esc(person.maxWeeklyLoad ?? 1)}" data-max-person="${esc(person.name)}"></td>
          </tr>`).join("")}</tbody>`;
      document.querySelectorAll("[data-max-person]").forEach((input) => {
        input.addEventListener("change", () => {
          const person = state.people.find((item) => item.name === input.dataset.maxPerson);
          if (!person) return;
          person.maxWeeklyLoad = Math.max(0, num(input.value, 1));
          renderAll();
        });
      });
    }
    function renderLines() {
      const rows = state.data.lineMetrics || [];
      document.getElementById("lineTable").innerHTML = `
        <thead><tr><th>Line</th><th class="right">Peak Demand</th><th>Peak Period</th><th class="right">Available</th><th class="right">Gap</th><th class="right">TBC Rows</th><th>Driver</th></tr></thead>
        <tbody>${rows.map((row) => `
          <tr>
            <td>${esc(row.line)}</td><td class="right">${fmt(row.peakDemand)}</td><td>${esc(row.peakMonth)}</td><td class="right">${fmt(row.available)}</td><td class="right">${fmt(row.gap)}</td><td class="right">${fmt(row.tbcRows, 0)}</td><td>${esc(row.why)}</td>
          </tr>`).join("")}</tbody>`;
    }
    function renderMatrix() {
      const rows = state.data.testLineMatrix || [];
      const lines = state.data.lines || [];
      document.getElementById("matrixTable").innerHTML = `
        <thead><tr><th>Test Type</th>${lines.map((line) => `<th class="right">${esc(line)}</th>`).join("")}<th class="right">Total Peak</th></tr></thead>
        <tbody>${rows.map((row) => `
          <tr><td>${esc(row.testType)}</td>${lines.map((line) => `<td class="right">${fmt(row[line])}</td>`).join("")}<td class="right">${fmt(row.totalPeak)}</td></tr>`).join("")}</tbody>`;
    }
    function renderAll() {
      derived = computeDerived();
      renderKpis();
      renderDemandChart();
      renderNotes();
      renderIssues();
      renderEditor();
      renderPeople();
      renderLines();
      renderMatrix();
    }
    function applyIssue(issue) {
      if (!issue || !issue.candidate) return;
      if (!canApplyCandidate(issue, issue.candidate, derived.loadIndex)) {
        alert("This recommendation affects other loaded periods. Please review and edit the allocation row manually.");
        return;
      }
      let remaining = num(issue.needed || issue.excessLoad || issue.load);
      const rows = rowsForIssueApply(issue);
      for (const row of rows) {
        if (remaining <= 1e-9) break;
        row.assigned = issue.candidate;
        const person = peopleByName().get(personKey(issue.candidate));
        row.company = clean(person?.company || "Unknown");
        remaining -= Math.max(rowPeakLoad(row, issue.period), 0.1);
      }
      renderAll();
    }
    function saveEdits() {
      const assignments = {};
      for (const row of state.rows) assignments[row.rowId] = { assigned: row.assigned, company: row.company };
      const capacities = {};
      for (const person of state.people) capacities[person.name] = person.maxWeeklyLoad;
      localStorage.setItem(STORAGE_KEY, JSON.stringify({ assignments, capacities }));
    }
    function loadEdits() {
      try {
        const saved = JSON.parse(localStorage.getItem(STORAGE_KEY) || "{}");
        if (saved.assignments) {
          for (const row of state.rows) {
            if (saved.assignments[row.rowId]) Object.assign(row, saved.assignments[row.rowId]);
          }
        }
        if (saved.capacities) {
          for (const person of state.people) {
            if (saved.capacities[person.name] !== undefined) person.maxWeeklyLoad = saved.capacities[person.name];
          }
        }
      } catch (error) {
        console.warn("Could not load saved dashboard edits", error);
      }
    }
    function resetEdits() {
      localStorage.removeItem(STORAGE_KEY);
      state.rows = clone(INITIAL_DATA.detailRows || []);
      state.people = clone(INITIAL_DATA.peopleRoster || []);
      renderAll();
    }
    function toCsv(rows, columns) {
      const quote = (value) => `"${String(value ?? "").replace(/"/g, '""')}"`;
      return [columns.map(quote).join(","), ...rows.map((row) => columns.map((col) => quote(row[col])).join(","))].join("\n");
    }
    function download(name, text, type = "text/csv") {
      const blob = new Blob([text], { type });
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = name;
      a.click();
      URL.revokeObjectURL(url);
    }
    function downloadIssues() {
      const rows = derived.issues.map((issue) => ({
        status: issue.status,
        period: issue.label,
        assigned: issue.assigned,
        line: issue.line || issue.lines || "",
        role: issue.role,
        load: fmt(issue.load, 2),
        threshold: fmt(issue.threshold, 2),
        excessLoad: fmt(issue.excessLoad, 2),
        suggestedFix: issue.suggestedFix
      }));
      download("resource_allocation_issues.csv", toCsv(rows, ["status", "period", "assigned", "line", "role", "load", "threshold", "excessLoad", "suggestedFix"]));
    }
    function downloadAllocation() {
      const rows = state.rows.map((row) => ({
        rowId: row.rowId,
        sourceRow: row.sourceRow,
        status: derived.rowStates.get(row.rowId) || "Available",
        line: row.line,
        testType: row.testType,
        activity: row.activity,
        role: row.role,
        assigned: row.assigned,
        company: row.company,
        start: row.start,
        finish: row.finish,
        requirement: row.requirement
      }));
      download("edited_resource_allocation.csv", toCsv(rows, ["rowId", "sourceRow", "status", "line", "testType", "activity", "role", "assigned", "company", "start", "finish", "requirement"]));
    }
    function initControls() {
      document.getElementById("sourceMeta").textContent = `${state.data.summary?.source || "Workbook"} - ${state.data.summary?.sourceMode || ""}`;
      const lineFilter = document.getElementById("lineFilter");
      lineFilter.innerHTML = ["All", ...(state.data.lines || [])].map((line) => `<option>${esc(line)}</option>`).join("");
      document.getElementById("viewMode").value = state.view;
      lineFilter.addEventListener("change", () => { state.line = lineFilter.value; renderAll(); });
      document.getElementById("statusFilter").addEventListener("change", (event) => { state.status = event.target.value; renderAll(); });
      document.getElementById("viewMode").addEventListener("change", (event) => { state.view = event.target.value; renderAll(); });
      document.getElementById("searchBox").addEventListener("input", (event) => { state.search = event.target.value; renderAll(); });
      document.getElementById("rowLimit").addEventListener("change", (event) => { state.limit = num(event.target.value, 250); renderAll(); });
      document.getElementById("saveBtn").addEventListener("click", saveEdits);
      document.getElementById("resetBtn").addEventListener("click", resetEdits);
      document.getElementById("issueCsvBtn").addEventListener("click", downloadIssues);
      document.getElementById("allocationCsvBtn").addEventListener("click", downloadAllocation);
      document.querySelectorAll("[data-tab]").forEach((button) => {
        button.addEventListener("click", () => {
          state.activeTab = button.dataset.tab;
          document.querySelectorAll(".tab-btn").forEach((btn) => btn.classList.toggle("active", btn.dataset.tab === state.activeTab));
          document.querySelectorAll(".panel").forEach((panel) => panel.classList.toggle("active", panel.id === `panel-${state.activeTab}`));
        });
      });
    }
    loadEdits();
    initControls();
    renderAll();
  </script>
</body>
</html>"""


# ==========================================
# STREAMLIT DASHBOARD UI
# ==========================================
def run_streamlit_app():
    if st is None or px is None:
        raise RuntimeError("Streamlit and Plotly are not installed. Use --static to generate the standalone HTML dashboard.")
    st.set_page_config(page_title="T&C Manpower Allocation", layout="wide")

    # --- 1. SIDEBAR CONTROLS ---
    st.sidebar.title("T&C Manpower Allocation")
    st.sidebar.caption("Line demand, test basis, and core-team floor.")
    
    st.sidebar.divider()
    uploaded_file = st.sidebar.file_uploader("Upload Excel Workbook", type=["xlsx", "xlsm", "xls"])
    
    # Process Data
    if uploaded_file:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
            
        try:
            DATA = build_data(workbook_path=tmp_path, source_label=uploaded_file.name)
        except Exception as e:
            st.sidebar.error(f"Error parsing workbook: {e}")
            st.stop()
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        DATA = empty_data()
        st.info("Please upload a T&C Resource Planning Report to populate the dashboard.")
        st.stop()

    # --- 2. FILTERS ---
    summary = DATA.get("summary", {})
    st.sidebar.divider()
    
    available_lines = DATA.get("lines", ["DTL", "JRL", "CRL", "RTS"])
    selected_lines = st.sidebar.multiselect("Line", available_lines, default=available_lines)
    
    test_types = ["All"] + sorted(DATA.get("testTypes", []))
    selected_test_type = st.sidebar.selectbox("Test Type", test_types, index=0)
    
    view_mode = st.sidebar.radio("View", ["Month", "Week"], index=0)
    
    if view_mode == "Week" and "Week" in DATA.get("demandRowsByView", {}):
        demand_rows = DATA["demandRowsByView"]["Week"]
        period_label = "weekLabel"
    else:
        demand_rows = DATA.get("demandRows", [])
        period_label = "monthLabel"

    # --- 3. TOP BAR & KPIs ---
    col1, col2 = st.columns([3, 1])
    with col1:
        st.subheader("Management View")
        st.caption(f"{summary.get('sourceMode', '')}: {summary.get('resourceRows', 0)} counted resource rows, " 
                   f"peak workbook demand {rounded(summary.get('peakDemand', 0))} in {summary.get('peakMonth', '')}.")
    with col2:
        excel_data = export_results_workbook(DATA)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name="Manpower_Allocation_Export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    filtered_detail = [r for r in DATA.get("detailRows", []) if r.get("line") in selected_lines]
    if selected_test_type != "All":
        filtered_detail = [r for r in filtered_detail if r.get("testType") == selected_test_type]

    tbc_count = sum(1 for r in filtered_detail if r.get("tbc") == "YES")
    active_types = len(set(r.get("testType") for r in filtered_detail))
    available_master = sum(DATA.get("manpower", {}).get(line, {}).get("total", 0) for line in selected_lines)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Peak Demand", rounded(summary.get("peakDemand", 0)), summary.get("peakMonth", ""))
    k2.metric("Available Master", available_master, ", ".join(selected_lines))
    k3.metric("Peak Gap", rounded(summary.get("peakDemand", 0) - available_master))
    k4.metric("TBC Rows", tbc_count, f"{summary.get('tbcRows', 0)} total")
    k5.metric("Test Types", active_types, "Filtered")
    k6.metric("Clash Periods", summary.get("overloadPersonMonths", 0))

    st.divider()

    # --- 4. CHARTS (Demand by Line) ---
    st.subheader("Demand by Line")
    if demand_rows:
        df_demand = pd.DataFrame(demand_rows)
        plot_lines = [line for line in selected_lines if line in df_demand.columns]
        
        if plot_lines:
            fig = px.bar(
                df_demand, 
                x="label", 
                y=plot_lines, 
                color_discrete_map=DATA.get("lineColors", {}),
                title="Stacked allocation vs available manpower",
                labels={"label": "Period", "value": "Load", "variable": "Line"}
            )
            
            threshold = DATA.get("summary", {}).get("totalManpower", 0)
            fig.add_hline(y=threshold, line_dash="dash", line_color="red", annotation_text="Max Capacity")
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No lines selected for chart.")
    else:
        st.info("No demand data available.")

    st.divider()

    # --- 5. DATA TABLES ---
    tab1, tab2, tab3, tab4 = st.tabs(["Line Summary", "Clash Detection", "Test Type Matrix", "Allocation Detail"])

    with tab1:
        st.subheader("Line Summary")
        metrics = [m for m in DATA.get("lineMetrics", []) if m.get("line") in selected_lines]
        if metrics:
            df_metrics = pd.DataFrame(metrics)[["line", "peakDemand", "peakMonth", "available", "gap", "tbcRows"]]
            st.dataframe(df_metrics, use_container_width=True, hide_index=True)
        else:
            st.write("No line data.")

    with tab2:
        st.subheader("Clash Detection")
        clashes = DATA.get("clashRows", [])
        if clashes:
            df_clashes = pd.DataFrame(clashes)[["assigned", period_label, "load", "excessLoad", "suggestedFix"]]
            st.dataframe(df_clashes.style.highlight_max(subset=['load'], color='#ff4b4b'), use_container_width=True, hide_index=True)
        else:
            st.success("No clashes detected! No named person exceeds 1.0 load.")

    with tab3:
        st.subheader("Test Type Comparison (Peak Load)")
        matrix = DATA.get("testLineMatrix", [])
        if matrix:
            df_matrix = pd.DataFrame(matrix)
            st.dataframe(df_matrix, use_container_width=True, hide_index=True)

    with tab4:
        st.subheader("Allocation Detail")
        if filtered_detail:
            df_detail = pd.DataFrame(filtered_detail)[
                ["line", "testType", "activity", "role", "assigned", "company", "tbc", "status"]
            ]
            
            def color_status(val):
                if val == 'Clash': return 'background-color: #ffcccc'
                elif val == 'TBC Gap': return 'background-color: #ffebcc'
                return ''
                
            st.dataframe(df_detail.style.map(color_status, subset=['status']), use_container_width=True, hide_index=True)
        else:
            st.write("No detail rows match your filters.")

def parse_args():
    parser = argparse.ArgumentParser(description="Build T&C resource allocation dashboard outputs.")
    parser.add_argument(
        "--workbook",
        default=str(Path.home() / "Downloads" / "Resource Planning 1-ET updates.xlsx"),
        help="Path to the resource planning workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(APP_DIR.parent / "outputs"),
        help="Directory for generated dashboard/report outputs.",
    )
    parser.add_argument(
        "--streamlit",
        action="store_true",
        help="Launch the Streamlit app instead of generating static outputs.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    if args.streamlit:
        run_streamlit_app()
        return
    outputs = generate_outputs(args.workbook, args.output_dir)
    summary = outputs["payload"].get("summary", {})
    print(f"Dashboard: {outputs['html']}")
    print(f"Excel report: {outputs['report']}")
    print(f"Patched code: {outputs['code']}")
    print(
        "Summary: "
        f"{summary.get('resourceRows', 0)} allocation rows, "
        f"{summary.get('personClashes', 0)} named clashes, "
        f"{summary.get('tbcGaps', 0)} TBC gaps."
    )


if __name__ == "__main__":
    main()
